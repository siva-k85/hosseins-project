"""
Enhanced eMMA -> Excel updater

Complete implementation with all requested features:
- Parameter validation with upfront checks
- argparse CLI implementation with subcommands
- Named logger configuration
- Dynamic backoff for 403/429 responses
- Enhanced error handling for workbook operations
- HTML schema resilience with header-based matching
- Additional fields capture
- Alternate timestamp formats support
- Duplicate detection
- Adaptive rate control
- Refs tab integration for auto-tagging
- Archive deduplication
- Cross-platform path handling
- Structured JSON logging
- Historical analytics
- Multi-date range scraping
- CLI subcommands
"""

import argparse
import os
import re
import time
import logging
import json
import csv
from hashlib import blake2b
from dataclasses import dataclass
from datetime import datetime, timedelta
from urllib.parse import urljoin
from zipfile import BadZipFile
from typing import Optional, Tuple, List, Dict, Any
import shutil
from pathlib import Path

import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

# -------------------- Config (env-overridable) --------------------

def validate_parameters():
    """Validate configuration parameters with helpful error messages."""
    errors = []

    # Validate DAYS_AGO
    days_ago = os.getenv("DAYS_AGO", "0")
    try:
        days_ago_int = int(days_ago)
        if days_ago_int < 0:
            errors.append(f"DAYS_AGO must be non-negative, got {days_ago_int}")
    except ValueError:
        errors.append(f"DAYS_AGO must be an integer, got '{days_ago}'")

    # Validate STALE_AFTER_D
    stale_after = os.getenv("STALE_AFTER_D", "7")
    try:
        stale_after_int = int(stale_after)
        if stale_after_int <= 0:
            errors.append(f"STALE_AFTER_D must be positive, got {stale_after_int}")
    except ValueError:
        errors.append(f"STALE_AFTER_D must be an integer, got '{stale_after}'")

    # Validate MAX_PAGES
    max_pages = os.getenv("MAX_PAGES", "50")
    try:
        max_pages_int = int(max_pages)
        if max_pages_int <= 0:
            errors.append(f"MAX_PAGES must be positive, got {max_pages_int}")
    except ValueError:
        errors.append(f"MAX_PAGES must be an integer, got '{max_pages}'")

    # Validate SLEEP_BETWEEN
    sleep_between = os.getenv("SLEEP_BETWEEN", "1.0")
    try:
        sleep_between_float = float(sleep_between)
        if sleep_between_float < 0:
            errors.append(f"SLEEP_BETWEEN must be non-negative, got {sleep_between_float}")
    except ValueError:
        errors.append(f"SLEEP_BETWEEN must be a number, got '{sleep_between}'")

    # Validate TIMEOUT_SECONDS
    timeout = os.getenv("TIMEOUT_SECONDS", "30")
    try:
        timeout_int = int(timeout)
        if timeout_int <= 0:
            errors.append(f"TIMEOUT_SECONDS must be positive, got {timeout_int}")
    except ValueError:
        errors.append(f"TIMEOUT_SECONDS must be an integer, got '{timeout}'")

    # Validate LOG_LEVEL
    log_level = os.getenv("LOG_LEVEL", "INFO").upper()
    valid_levels = ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"]
    if log_level not in valid_levels:
        errors.append(f"LOG_LEVEL must be one of {valid_levels}, got '{log_level}'")

    if errors:
        raise ValueError("Configuration validation failed:\n" + "\n".join(f"  - {e}" for e in errors))

# Cross-platform path handling
def get_default_workbook_path():
    """Get platform-appropriate default workbook path."""
    if os.name == 'nt':  # Windows
        return r"C:\Users\hkhoshhal001\Guidehouse\New York Mid-Atlantic (NYMA) - 05. Scanning resources\Automated Scanning\opportunities.xlsx"
    else:  # Unix/Linux/Mac
        home = os.path.expanduser("~")
        return os.path.join(home, "Documents", "emma_opportunities.xlsx")

WORKBOOK_PATH = os.getenv("EMMA_XLSX", get_default_workbook_path())

# Validate parameters before using them
try:
    validate_parameters()
except ValueError as e:
    # Log error but continue with defaults for now
    print(f"Warning: {e}")

DAYS_AGO        = int(os.getenv("DAYS_AGO", "0"))
STALE_AFTER_D   = int(os.getenv("STALE_AFTER_D", "7"))
MAX_PAGES       = int(os.getenv("MAX_PAGES", "50"))
SLEEP_BETWEEN   = float(os.getenv("SLEEP_BETWEEN", "1.0"))
LOG_LEVEL       = os.getenv("LOG_LEVEL", "INFO").upper()
TIMEOUT_SECONDS = int(os.getenv("TIMEOUT_SECONDS", "30"))
USER_AGENT      = os.getenv("USER_AGENT", "Mozilla/5.0 (compatible; MD-EmmaScraper/1.0)")

BASE = "https://emma.maryland.gov"
BROWSE_URL = f"{BASE}/page.aspx/en/rfp/request_browse_public"

# -------------------- Timezone (ET) --------------------
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    ET_TZ = ZoneInfo("America/New_York")
except Exception:
    ET_TZ = None

def now_et():
    return datetime.now(ET_TZ) if ET_TZ else datetime.now()

def localize_et(dt: datetime) -> datetime:
    return dt.replace(tzinfo=ET_TZ) if ET_TZ else dt

def to_excel_naive(dt):
    """Excel/openpyxl cannot store tz-aware datetimes. Return naive (no tzinfo)."""
    if dt is None:
        return None
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt

# -------------------- Logging --------------------
# Create a named logger
logger = logging.getLogger("emma_scraper")

class JsonFormatter(logging.Formatter):
    """JSON formatter for structured logging."""
    def format(self, record):
        log_obj = {
            "timestamp": datetime.fromtimestamp(record.created).isoformat(),
            "level": record.levelname,
            "logger": record.name,
            "message": record.getMessage(),
            "module": record.module,
            "function": record.funcName,
            "line": record.lineno
        }
        if record.exc_info:
            log_obj["exception"] = self.formatException(record.exc_info)
        if hasattr(record, 'extra_data'):
            log_obj["data"] = record.extra_data
        return json.dumps(log_obj)

def configure_logging(level: str, json_format: bool = False):
    """Configure the named logger with appropriate handlers."""
    logger.setLevel(getattr(logging, level, logging.INFO))

    # Clear any existing handlers
    logger.handlers.clear()

    # Create console handler
    console_handler = logging.StreamHandler()

    if json_format:
        console_handler.setFormatter(JsonFormatter())
    else:
        # Standard text format
        formatter = logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S"
        )
        console_handler.setFormatter(formatter)

    logger.addHandler(console_handler)

    # Reduce noise from other libraries
    logging.getLogger("requests").setLevel(logging.WARNING)
    logging.getLogger("urllib3").setLevel(logging.WARNING)

# -------------------- Excel Styling Module --------------------
class ExcelStyler:
    """Centralized Excel styling module."""

    @staticmethod
    def get_status_fill(status: str) -> PatternFill:
        """Get fill color for status."""
        fills = {
            "New": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "Updated": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "Unchanged": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "Stale": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        }
        return fills.get(status, PatternFill())

    @staticmethod
    def apply_header_style(ws):
        """Apply header styling."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    @staticmethod
    def apply_borders(ws):
        """Apply borders to all cells with data."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

# -------------------- Enhanced HTTP Session with Dynamic Backoff --------------------
class DynamicRetry(Retry):
    """Custom retry with dynamic backoff for 403/429 responses."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.retry_after_header = None

    def get_backoff_time(self):
        """Get backoff time, respecting Retry-After header if present."""
        if self.retry_after_header:
            try:
                return int(self.retry_after_header)
            except ValueError:
                pass
        return super().get_backoff_time()

    def increment(self, method=None, url=None, response=None, error=None, _pool=None, _stacktrace=None):
        """Increment retry count and check for Retry-After header."""
        if response:
            self.retry_after_header = response.headers.get('Retry-After')

            # Dynamic backoff for rate limiting
            if response.status_code in [403, 429]:
                logger.warning(f"Rate limited (HTTP {response.status_code}). Implementing dynamic backoff.")
                if self.retry_after_header:
                    logger.info(f"Server requested retry after {self.retry_after_header} seconds")

        return super().increment(method, url, response, error, _pool, _stacktrace)

def make_session() -> requests.Session:
    """Create HTTP session with enhanced retry logic."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    })

    retries = DynamicRetry(
        total=5,
        backoff_factor=1.0,
        status_forcelist=[403, 429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        raise_on_status=False
    )

    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.mount("http://", HTTPAdapter(max_retries=retries))
    return s

# -------------------- Enhanced eMMA Scraping --------------------
# Enhanced timestamp patterns and formats
TS_PATTERN = re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s+(AM|PM)\b")

# Extended timestamp formats for better parsing
DETAIL_TIMESTAMP_FORMATS = [
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%m-%d-%Y %I:%M:%S %p",
    "%m-%d-%Y %I:%M %p",
    "%m-%d-%Y",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y",
]

PUBLISH_DT_FORMATS = DETAIL_TIMESTAMP_FORMATS

# Extended field labels for better extraction
DETAIL_SUMMARY_LABELS = ["summary", "description", "project description", "scope", "overview", "details"]
DETAIL_OFFICER_LABELS = ["procurement officer", "buyer", "contact", "procurement contact", "issuing officer", "contact person"]
DETAIL_EMAIL_LABELS = ["email", "e-mail", "contact email", "buyer email", "officer email"]
DETAIL_INSTRUCTIONS_LABELS = ["instruction", "guideline", "submission", "note", "requirement", "special instruction", "how to"]
DETAIL_GOALS_LABELS = ["goal", "participation", "mbe", "dbe", "sbe", "wbe", "small business", "program goal", "diversity"]
DETAIL_DUE_LABELS = ["due date", "bid due date", "proposal due date", "response due date", "closing date", "closing time", "deadline"]
DETAIL_SOLICITATION_LABELS = ["solicitation id", "solicitation #", "solicitation number", "rfp #", "bid #", "contract #", "reference #"]
DETAIL_CONTACT_PHONE_LABELS = ["phone", "telephone", "tel", "contact number", "phone number"]

DETAIL_FIELD_KEYS = [
    "solicitation_summary",
    "procurement_officer_buyer",
    "contact_email",
    "contact_phone",
    "additional_instructions",
    "procurement_program_goals",
]

EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_REGEX = re.compile(r"[\d\s\(\)\-\+\.]{10,}")

# Enhanced header aliases for resilient matching
LISTING_HEADER_ALIASES = {
    "title": {
        "title", "solicitation", "solicitation title", "project title", "opportunity",
        "description", "project", "name", "subject"
    },
    "solicitation_id": {
        "solicitation id", "solicitation #", "solicitation number", "reference #", "document number",
        "document id", "document #", "rfp number", "bid number", "contract number",
        "id", "ref", "reference", "tracking #", "tracking number"
    },
    "category": {
        "category", "procurement category", "commodity", "product category",
        "type", "classification", "commodity category"
    },
    "procurement_method": {
        "procurement method", "method", "contract type", "procurement type",
        "bid type", "solicitation type", "acquisition method"
    },
    "agency": {
        "agency", "issuing agency", "department", "agency/department",
        "organization", "entity", "buyer", "issuing department"
    },
    "publish_dt": {
        "publish date", "posting date", "posted", "issue date", "posting date/time",
        "published", "date posted", "release date", "available date"
    },
    "due_dt": {
        "due date", "bid due date", "proposal due date", "response due date",
        "closing date", "closing time", "deadline", "submission deadline",
        "due", "closes", "bid close"
    },
}

# Excel column headers
MASTER_HDR = [
    "source","record_id","url","first_seen_et","last_seen_et",
    "title","agency","category","procurement_method","publish_dt_et","due_dt_et","solicitation_id",
    "solicitation_summary","procurement_officer_buyer","contact_email","contact_phone",
    "additional_instructions","procurement_program_goals",
    "status","tags","score_bd_fit"
]

def parse_hidden_fields(soup: BeautifulSoup) -> dict:
    """Parse ASP.NET hidden fields."""
    fields = {}
    for name in ["__VIEWSTATE","__EVENTVALIDATION","__VIEWSTATEGENERATOR","__EVENTTARGET","__EVENTARGUMENT"]:
        el = soup.find("input", {"name": name})
        if el and el.has_attr("value"):
            fields[name] = el["value"]
    return fields

def _normalize_text(text: str) -> str:
    """Normalize text by collapsing whitespace."""
    return re.sub(r"\s+", " ", (text or "").strip())

def _build_column_map(header_cells) -> dict:
    """Build column mapping with enhanced resilience."""
    mapping = {}
    for idx, cell in enumerate(header_cells or []):
        label = _normalize_text(cell.get_text(" ", strip=True)).lower()
        if not label:
            continue

        # Try exact match first
        for canonical, aliases in LISTING_HEADER_ALIASES.items():
            if label in aliases:
                mapping[canonical] = idx
                break

        # Try partial match if no exact match
        if canonical not in mapping:
            for canonical, aliases in LISTING_HEADER_ALIASES.items():
                if any(alias in label for alias in aliases):
                    mapping[canonical] = idx
                    break

    return mapping

def parse_flexible_datetime(date_str: str) -> Optional[datetime]:
    """Parse datetime with multiple format support."""
    if not date_str:
        return None

    cleaned = _normalize_text(date_str)

    # Try all formats
    for fmt in DETAIL_TIMESTAMP_FORMATS:
        try:
            dt = datetime.strptime(cleaned, fmt)
            return localize_et(dt)
        except ValueError:
            continue

    # Try to extract date pattern with regex
    match = TS_PATTERN.search(cleaned)
    if match:
        for fmt in DETAIL_TIMESTAMP_FORMATS:
            try:
                dt = datetime.strptime(match.group(0), fmt)
                return localize_et(dt)
            except ValueError:
                continue

    logger.debug(f"Failed to parse datetime: '{cleaned}'")
    return None

def extract_rows_enhanced(soup: BeautifulSoup) -> list[dict]:
    """Extract rows with enhanced header matching."""
    # Try multiple table selectors
    table = None
    for selector in ["table.iv-grid-view", "table.grid", "table#results", "table"]:
        table = soup.select_one(selector)
        if table:
            break

    if not table:
        logger.warning("No results table found on page")
        return []

    rows = []
    tbody = table.find("tbody")
    body_rows = tbody.find_all("tr") if tbody else table.find_all("tr")

    # Find header row
    header_row = None
    header_cells = []
    thead = table.find("thead")
    if thead:
        header_row = thead.find("tr")
    elif body_rows:
        first = body_rows[0]
        if first.find_all("th"):
            header_row = first

    if header_row:
        header_cells = header_row.find_all(["th", "td"])
        body_rows = [row for row in body_rows if row is not header_row]

    column_map = _build_column_map(header_cells)

    for tr in body_rows:
        tds = tr.find_all("td")
        if not tds:
            continue

        # Extract with fallbacks
        row_data = {}

        # Title with link
        title_idx = column_map.get("title", 0)
        if title_idx < len(tds):
            title_cell = tds[title_idx]
            a = title_cell.find("a")
            row_data["title"] = a.get_text(strip=True) if a else title_cell.get_text(strip=True)
            row_data["url"] = urljoin(BASE, a["href"]) if (a and a.has_attr("href")) else None

        # Other fields
        for field in ["category", "procurement_method", "agency", "solicitation_id"]:
            idx = column_map.get(field)
            if idx is not None and idx < len(tds):
                row_data[field] = tds[idx].get_text(" ", strip=True)
            else:
                row_data[field] = ""

        # Dates with flexible parsing
        for date_field in ["publish_dt", "due_dt"]:
            idx = column_map.get(date_field)
            if idx is not None and idx < len(tds):
                date_text = tds[idx].get_text(" ", strip=True)
                row_data[f"{date_field}_raw"] = date_text
            else:
                row_data[f"{date_field}_raw"] = ""

        rows.append(row_data)

    return rows

def scrape_detail_enhanced(session: requests.Session, url: str) -> dict:
    """Enhanced detail page scraping with more fields."""
    detail_payload = {key: "" for key in DETAIL_FIELD_KEYS}
    detail_payload.update({
        "detail_due_text": "",
        "due_dt_et": None,
        "__fetched__": False,
    })

    if not url:
        return detail_payload

    try:
        response = session.get(url, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
    except Exception as exc:
        logger.warning(f"Failed to fetch detail page {url}: {exc}")
        return detail_payload

    soup = BeautifulSoup(response.text, "html.parser")

    # Extract all labeled values
    labeled_values = {}
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) >= 2:
                label = _normalize_text(cells[0].get_text(" ", strip=True)).lower()
                value = _normalize_text(cells[1].get_text(" ", strip=True))
                if label and value:
                    labeled_values[label] = value

    # Extract fields with fallbacks
    for label_set, field_key in [
        (DETAIL_SUMMARY_LABELS, "solicitation_summary"),
        (DETAIL_OFFICER_LABELS, "procurement_officer_buyer"),
        (DETAIL_INSTRUCTIONS_LABELS, "additional_instructions"),
        (DETAIL_GOALS_LABELS, "procurement_program_goals"),
    ]:
        for label in label_set:
            if any(label in k for k in labeled_values.keys()):
                matching_key = next(k for k in labeled_values.keys() if label in k)
                detail_payload[field_key] = labeled_values[matching_key]
                break

    # Extract email with regex
    email_value = ""
    for label in DETAIL_EMAIL_LABELS:
        if any(label in k for k in labeled_values.keys()):
            matching_key = next(k for k in labeled_values.keys() if label in k)
            email_value = labeled_values[matching_key]
            break

    if email_value:
        match = EMAIL_REGEX.search(email_value)
        if match:
            detail_payload["contact_email"] = match.group(0)
    else:
        # Search entire page
        match = EMAIL_REGEX.search(soup.get_text())
        if match:
            detail_payload["contact_email"] = match.group(0)

    # Extract phone
    phone_value = ""
    for label in DETAIL_CONTACT_PHONE_LABELS:
        if any(label in k for k in labeled_values.keys()):
            matching_key = next(k for k in labeled_values.keys() if label in k)
            phone_value = labeled_values[matching_key]
            break

    if phone_value:
        match = PHONE_REGEX.search(phone_value)
        if match:
            detail_payload["contact_phone"] = match.group(0).strip()

    # Extract due date
    for label in DETAIL_DUE_LABELS:
        if any(label in k for k in labeled_values.keys()):
            matching_key = next(k for k in labeled_values.keys() if label in k)
            due_text = labeled_values[matching_key]
            detail_payload["detail_due_text"] = due_text
            detail_payload["due_dt_et"] = parse_flexible_datetime(due_text)
            break

    detail_payload["__fetched__"] = True
    return detail_payload

def emma_scrape_enhanced(days_ago: int, max_pages: int = 50, sleep_s: float = 1.0,
                         fetch_details: bool = True, adaptive_rate: bool = True) -> list[dict]:
    """Enhanced scraping with adaptive rate control and duplicate detection."""
    ses = make_session()
    r = ses.get(BROWSE_URL, timeout=TIMEOUT_SECONDS)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    all_rows = []
    pages = 0
    seen_fingerprints = set()
    current_sleep = sleep_s

    def page_fingerprint(s):
        """Create page fingerprint to detect duplicates."""
        t = s.select_one("table") or s
        text = t.get_text(" ", strip=True)[:20000]
        return blake2b(text.encode(), digest_size=16).hexdigest()

    while True:
        fp = page_fingerprint(soup)
        if fp in seen_fingerprints:
            logger.info("Detected duplicate page, stopping pagination")
            break
        seen_fingerprints.add(fp)

        rows = extract_rows_enhanced(soup)
        all_rows.extend(rows)
        pages += 1

        logger.info(f"Scraped page {pages}, found {len(rows)} rows")

        if pages >= max_pages:
            logger.info(f"Reached max pages limit ({max_pages})")
            break

        # Find next page
        et, ea = find_next_postback(soup)
        if not et:
            logger.info("No next page found")
            break

        # Adaptive rate control
        if adaptive_rate and r.headers.get('X-RateLimit-Remaining'):
            try:
                remaining = int(r.headers['X-RateLimit-Remaining'])
                if remaining < 10:
                    current_sleep = min(current_sleep * 2, 10)
                    logger.info(f"Rate limit low ({remaining}), increasing sleep to {current_sleep}s")
            except ValueError:
                pass

        fields = parse_hidden_fields(soup)
        fields["__EVENTTARGET"] = et
        fields["__EVENTARGUMENT"] = ea

        time.sleep(current_sleep)

        r = ses.post(BROWSE_URL, data=fields, timeout=TIMEOUT_SECONDS)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

    # Fetch details
    if fetch_details and all_rows:
        success = 0
        failure = 0
        total = len(all_rows)

        for idx, row in enumerate(all_rows, start=1):
            detail = scrape_detail_enhanced(ses, row.get("url"))

            if detail.get("__fetched__"):
                success += 1
            else:
                failure += 1

            # Merge detail fields
            for key in DETAIL_FIELD_KEYS:
                row[key] = detail.get(key, "")

            if detail.get("due_dt_et") and not row.get("due_dt_et"):
                row["due_dt_et"] = detail["due_dt_et"]

            if idx % 10 == 0 or idx == total:
                logger.info(f"Fetched {idx}/{total} detail pages...")

            if idx < total:
                time.sleep(current_sleep)

        logger.info(f"Detail fetch complete: {success} success, {failure} failed")

    # Process and normalize rows
    for row in all_rows:
        # Parse dates
        publish_raw = row.pop("publish_dt_raw", "")
        row["publish_dt_et"] = parse_flexible_datetime(publish_raw)

        due_raw = row.pop("due_dt_raw", "")
        if not row.get("due_dt_et"):
            row["due_dt_et"] = parse_flexible_datetime(due_raw)

        # Set defaults
        row["source"] = "emma"
        row["solicitation_id"] = (row.get("solicitation_id") or "").strip()
        row.setdefault("tags", "")
        row.setdefault("score_bd_fit", "")

        # Create record ID
        row["record_id"] = make_record_id_enhanced(row)

    # Deduplicate using title+agency+publish_dt
    deduped = deduplicate_rows_enhanced(all_rows)

    # Filter by target date
    target_date = (now_et().date() - timedelta(days=days_ago))
    staging = [r for r in deduped if r.get("publish_dt_et") and r["publish_dt_et"].date() == target_date]

    logger.info(f"Found {len(staging)} records for {target_date}")

    return staging

def make_record_id_enhanced(row: dict) -> str:
    """Create stable record ID with multiple fallbacks."""
    # Try solicitation ID first
    solicitation_id = (row.get("solicitation_id") or "").strip()
    if solicitation_id:
        return f"sid_{solicitation_id}"

    # Try URL-based ID
    url = row.get("url") or ""
    if url:
        match = re.search(r"/extranet/(\d+)", url)
        if match:
            return f"eid_{match.group(1)}"

    # Create composite key from title+agency+publish_dt
    title = (row.get("title") or "").strip()
    agency = (row.get("agency") or "").strip()
    publish_dt = row.get("publish_dt_et")

    if title and agency and publish_dt:
        seed = f"{title}|{agency}|{publish_dt.isoformat()}".encode()
        return f"hash_{blake2b(seed, digest_size=8).hexdigest()}"

    # Final fallback
    seed = (url + title).encode()
    return f"emma_{blake2b(seed, digest_size=8).hexdigest()}"

def deduplicate_rows_enhanced(rows: List[dict]) -> List[dict]:
    """Enhanced deduplication using composite keys."""
    deduped = []
    seen_keys = set()
    duplicate_count = 0

    for row in rows:
        # Create composite key
        key = (
            (row.get("title") or "").strip().lower(),
            (row.get("agency") or "").strip().lower(),
            row.get("publish_dt_et").isoformat() if row.get("publish_dt_et") else "",
            (row.get("solicitation_id") or "").strip().lower(),
        )

        if key in seen_keys:
            duplicate_count += 1
            logger.debug(f"Skipping duplicate: {key}")
            continue

        seen_keys.add(key)
        deduped.append(row)

    if duplicate_count:
        logger.info(f"Removed {duplicate_count} duplicate rows")

    return deduped

def find_next_postback(soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str]]:
    """Find next page postback parameters."""
    # Try explicit next link
    for a in soup.find_all("a", href=True):
        m = re.search(r"__doPostBack\('([^']*)','([^']*)'\)", a["href"])
        if m:
            text = a.get_text(strip=True).lower()
            if text in {"next", ">", ">>", "next page"}:
                return m.group(1), m.group(2)

    # Fallback to highest page number
    candidates = []
    for a in soup.find_all("a", href=True):
        m = re.search(r"__doPostBack\('([^']*)','([^']*)'\)", a["href"])
        label = a.get_text(strip=True)
        if m and label.isdigit():
            candidates.append((int(label), m.group(1), m.group(2)))

    if candidates:
        _, et, ea = sorted(candidates, key=lambda x: x[0])[-1]
        return et, ea

    return None, None

# -------------------- Refs Integration for Auto-Tagging --------------------
def load_refs_rules(wb) -> List[Dict[str, Any]]:
    """Load tagging rules from Refs sheet."""
    if "Refs" not in wb.sheetnames:
        return []

    ws_refs = wb["Refs"]
    if ws_refs.max_row < 2:
        return []

    rules = []
    header = [c.value for c in ws_refs[1]]

    if not header or len(header) < 3:
        # Initialize Refs header if missing
        ws_refs.delete_rows(1, ws_refs.max_row)
        ws_refs.append(["keyword", "field", "tag", "score", "priority"])
        return []

    for row_num in range(2, ws_refs.max_row + 1):
        row_data = [ws_refs.cell(row=row_num, column=i+1).value for i in range(len(header))]
        rule = dict(zip(header, row_data))
        if rule.get("keyword") and rule.get("tag"):
            rules.append(rule)

    # Sort by priority if available
    rules.sort(key=lambda x: x.get("priority", 999))

    return rules

def apply_auto_tagging(row: dict, rules: List[Dict[str, Any]]) -> dict:
    """Apply auto-tagging rules to a row."""
    tags = []
    score = 0

    for rule in rules:
        keyword = (rule.get("keyword") or "").lower()
        field = rule.get("field", "title")
        tag = rule.get("tag", "")
        rule_score = int(rule.get("score", 0))

        # Check if keyword matches in specified field
        field_value = (row.get(field) or "").lower()
        if keyword in field_value:
            if tag and tag not in tags:
                tags.append(tag)
            score += rule_score

    # Update row
    if tags:
        existing_tags = row.get("tags", "")
        if existing_tags:
            all_tags = existing_tags.split(",") + tags
        else:
            all_tags = tags
        row["tags"] = ",".join(list(set(all_tags)))

    if score > 0:
        existing_score = row.get("score_bd_fit", "")
        if existing_score and existing_score.isdigit():
            score += int(existing_score)
        row["score_bd_fit"] = str(score)

    return row

# -------------------- Archive Deduplication --------------------
def deduplicate_archive(wb):
    """Remove duplicate records from Archive sheet."""
    if "Archive" not in wb.sheetnames:
        return 0

    ws_archive = wb["Archive"]
    if ws_archive.max_row < 2:
        return 0

    header = [c.value for c in ws_archive[1]]
    if "record_id" not in header:
        return 0

    record_id_col = header.index("record_id") + 1

    # Find duplicates
    seen = set()
    rows_to_delete = []

    for row_num in range(2, ws_archive.max_row + 1):
        record_id = ws_archive.cell(row=row_num, column=record_id_col).value
        if record_id in seen:
            rows_to_delete.append(row_num)
        else:
            seen.add(record_id)

    # Delete duplicates (bottom-up)
    removed = 0
    for row_num in sorted(rows_to_delete, reverse=True):
        ws_archive.delete_rows(row_num, 1)
        removed += 1

    if removed:
        logger.info(f"Removed {removed} duplicate rows from Archive")

    return removed

# -------------------- Historical Analytics --------------------
def generate_analytics_report(wb, output_path: str, format_type: str = "csv"):
    """Generate analytics report from workbook data."""
    if "Master" not in wb.sheetnames:
        logger.error("Master sheet not found")
        return

    ws_master = wb["Master"]

    # Extract data
    data = []
    header = [c.value for c in ws_master[1]]

    for row_num in range(2, ws_master.max_row + 1):
        row_data = [ws_master.cell(row=row_num, column=i+1).value for i in range(len(header))]
        data.append(dict(zip(header, row_data)))

    # Calculate analytics
    analytics = {
        "total_records": len(data),
        "by_status": {},
        "by_agency": {},
        "by_category": {},
        "by_date": {},
        "avg_days_to_due": [],
    }

    for row in data:
        # Count by status
        status = row.get("status", "Unknown")
        analytics["by_status"][status] = analytics["by_status"].get(status, 0) + 1

        # Count by agency
        agency = row.get("agency", "Unknown")
        analytics["by_agency"][agency] = analytics["by_agency"].get(agency, 0) + 1

        # Count by category
        category = row.get("category", "Unknown")
        analytics["by_category"][category] = analytics["by_category"].get(category, 0) + 1

        # Count by publish date
        publish_dt = row.get("publish_dt_et")
        if publish_dt:
            date_key = publish_dt.date() if isinstance(publish_dt, datetime) else str(publish_dt)
            analytics["by_date"][str(date_key)] = analytics["by_date"].get(str(date_key), 0) + 1

        # Calculate days to due
        if row.get("publish_dt_et") and row.get("due_dt_et"):
            try:
                delta = (row["due_dt_et"] - row["publish_dt_et"]).days
                analytics["avg_days_to_due"].append(delta)
            except:
                pass

    # Calculate average
    if analytics["avg_days_to_due"]:
        avg_days = sum(analytics["avg_days_to_due"]) / len(analytics["avg_days_to_due"])
        analytics["avg_days_to_due"] = round(avg_days, 1)
    else:
        analytics["avg_days_to_due"] = 0

    # Export based on format
    if format_type == "json":
        with open(output_path, 'w') as f:
            json.dump(analytics, f, indent=2, default=str)

    elif format_type == "xlsx":
        wb_report = Workbook()
        ws = wb_report.active
        ws.title = "Analytics"

        # Summary section
        ws.append(["Metric", "Value"])
        ws.append(["Total Records", analytics["total_records"]])
        ws.append(["Avg Days to Due", analytics["avg_days_to_due"]])
        ws.append([])

        # By Status
        ws.append(["Status", "Count"])
        for status, count in analytics["by_status"].items():
            ws.append([status, count])
        ws.append([])

        # By Agency (top 10)
        ws.append(["Top Agencies", "Count"])
        for agency, count in sorted(analytics["by_agency"].items(), key=lambda x: x[1], reverse=True)[:10]:
            ws.append([agency, count])

        wb_report.save(output_path)

    else:  # CSV
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Total Records", analytics["total_records"]])
            writer.writerow(["Avg Days to Due", analytics["avg_days_to_due"]])
            writer.writerow([])

            writer.writerow(["Status", "Count"])
            for status, count in analytics["by_status"].items():
                writer.writerow([status, count])
            writer.writerow([])

            writer.writerow(["Agency", "Count"])
            for agency, count in sorted(analytics["by_agency"].items(), key=lambda x: x[1], reverse=True)[:10]:
                writer.writerow([agency, count])

    logger.info(f"Analytics report saved to {output_path}")

# -------------------- Enhanced Workbook Operations --------------------
def create_workbook_backup(original_path: str, max_backups: int = 5) -> str:
    """Create timestamped backup of workbook."""
    if not os.path.exists(original_path):
        return ""

    try:
        backup_dir = os.path.join(os.path.dirname(original_path), "backups")
        os.makedirs(backup_dir, exist_ok=True)

        base_name = os.path.splitext(os.path.basename(original_path))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{base_name}_backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)

        shutil.copy2(original_path, backup_path)
        logger.info(f"Created backup: {backup_path}")

        # Cleanup old backups
        cleanup_old_backups(backup_dir, base_name, max_backups)

        return backup_path

    except Exception as exc:
        logger.warning(f"Backup failed: {exc}")
        return ""

def cleanup_old_backups(backup_dir: str, base_name: str, max_backups: int):
    """Remove old backups beyond limit."""
    try:
        backups = []
        for f in os.listdir(backup_dir):
            if f.startswith(f"{base_name}_backup_") and f.endswith(".xlsx"):
                path = os.path.join(backup_dir, f)
                backups.append((path, os.path.getmtime(path)))

        backups.sort(key=lambda x: x[1], reverse=True)

        for path, _ in backups[max_backups:]:
            os.remove(path)
            logger.debug(f"Removed old backup: {path}")

    except Exception as exc:
        logger.warning(f"Cleanup failed: {exc}")

def load_workbook_safe(path: str):
    """Load workbook with enhanced error handling."""
    try:
        # Initialize if missing
        if not os.path.exists(path):
            init_workbook_enhanced(path)

        return load_workbook(path)

    except (BadZipFile, InvalidFileException) as exc:
        logger.warning(f"Workbook corrupted: {exc}. Reinitializing...")

        # Backup corrupted file
        if os.path.exists(path):
            corrupted_path = path + ".corrupted"
            shutil.move(path, corrupted_path)
            logger.info(f"Moved corrupted file to {corrupted_path}")

        init_workbook_enhanced(path)
        return load_workbook(path)

    except PermissionError as exc:
        logger.error(f"Permission denied: {exc}")
        raise RuntimeError(f"Cannot access workbook '{path}': Permission denied. Please close the file if it's open.") from exc

    except OSError as exc:
        logger.error(f"OS error: {exc}")
        raise RuntimeError(f"Cannot open workbook '{path}': {exc}") from exc

def init_workbook_enhanced(path: str):
    """Initialize workbook with all sheets."""
    wb = Workbook()

    # Master sheet
    ws_master = wb.active
    ws_master.title = "Master"
    ws_master.append(MASTER_HDR)

    # Log sheet
    ws_log = wb.create_sheet("Log")
    ws_log.append(["run_ts_et", "action"] + MASTER_HDR)

    # Archive sheet
    ws_archive = wb.create_sheet("Archive")
    ws_archive.append(MASTER_HDR)

    # Refs sheet for auto-tagging rules
    ws_refs = wb.create_sheet("Refs")
    ws_refs.append(["keyword", "field", "tag", "score", "priority", "notes"])

    # Add sample rules
    ws_refs.append(["construction", "title", "Construction", "10", "1", "Construction projects"])
    ws_refs.append(["it services", "title", "IT", "15", "2", "IT and technology services"])
    ws_refs.append(["consulting", "title", "Consulting", "8", "3", "Consulting services"])

    # Analytics sheet (optional)
    ws_analytics = wb.create_sheet("Analytics")
    ws_analytics.append(["Date", "New", "Updated", "Unchanged", "Stale", "Total"])

    wb.save(path)
    logger.info(f"Initialized workbook: {path}")

# -------------------- Main Merge Pipeline --------------------
@dataclass
class Action:
    action: str
    row: dict

def merge_into_excel_enhanced(staging: list[dict], wb_path: str):
    """Enhanced merge with all features."""
    ts_run = now_et()
    ts_run_xl = to_excel_naive(ts_run)

    # Load workbook with error handling
    wb = load_workbook_safe(wb_path)

    ws_master = wb["Master"]
    ws_log = wb["Log"]
    ws_archive = wb["Archive"]
    ws_analytics = wb.get("Analytics")

    # Load auto-tagging rules
    rules = load_refs_rules(wb)
    if rules:
        logger.info(f"Loaded {len(rules)} auto-tagging rules")

    # Build index
    header = [c.value for c in ws_master[1]]
    col_idx = {name: header.index(name)+1 for name in header}
    index = {}
    for row in range(2, ws_master.max_row+1):
        rid = ws_master.cell(row=row, column=col_idx["record_id"]).value
        if rid:
            index[rid] = row

    actions = []
    touched_ids = set()

    # Process staging rows
    for r in staging:
        # Apply auto-tagging
        if rules:
            r = apply_auto_tagging(r, rules)

        row = {
            "source": r.get("source", "emma"),
            "record_id": r["record_id"],
            "url": r.get("url"),
            "first_seen_et": None,
            "last_seen_et": ts_run_xl,
            "title": r.get("title", ""),
            "agency": r.get("agency", ""),
            "category": r.get("category", ""),
            "procurement_method": r.get("procurement_method", ""),
            "publish_dt_et": to_excel_naive(r.get("publish_dt_et")),
            "due_dt_et": to_excel_naive(r.get("due_dt_et")),
            "solicitation_id": r.get("solicitation_id", ""),
            "solicitation_summary": r.get("solicitation_summary", ""),
            "procurement_officer_buyer": r.get("procurement_officer_buyer", ""),
            "contact_email": r.get("contact_email", ""),
            "contact_phone": r.get("contact_phone", ""),
            "additional_instructions": r.get("additional_instructions", ""),
            "procurement_program_goals": r.get("procurement_program_goals", ""),
            "status": None,
            "tags": r.get("tags", ""),
            "score_bd_fit": r.get("score_bd_fit", ""),
        }

        rid = row["record_id"]
        touched_ids.add(rid)

        if rid not in index:
            # New record
            row["first_seen_et"] = ts_run_xl
            row["status"] = "New"
            ws_master.append([row[k] for k in MASTER_HDR])
            actions.append(Action("New", row))
            index[rid] = ws_master.max_row
        else:
            # Existing record
            existing_row = [ws_master.cell(row=index[rid], column=i+1).value for i in range(len(header))]
            existing = dict(zip(header, existing_row))

            row["first_seen_et"] = existing.get("first_seen_et") or ts_run_xl

            # Check if updated
            changed = False
            for k in ["title", "agency", "category", "url", "due_dt_et"]:
                if existing.get(k) != row.get(k):
                    changed = True
                    break

            if changed:
                # Update row
                for k in MASTER_HDR:
                    if k in col_idx:
                        ws_master.cell(row=index[rid], column=col_idx[k]).value = row[k]
                ws_master.cell(row=index[rid], column=col_idx["status"]).value = "Updated"
                actions.append(Action("Updated", row))
            else:
                # Unchanged
                ws_master.cell(row=index[rid], column=col_idx["last_seen_et"]).value = ts_run_xl
                ws_master.cell(row=index[rid], column=col_idx["status"]).value = "Unchanged"
                actions.append(Action("Unchanged", row))

    # Archive stale records
    stale_cutoff = to_excel_naive(ts_run - timedelta(days=STALE_AFTER_D))
    to_archive = []

    for rid, rownum in list(index.items()):
        if rid in touched_ids:
            continue

        last_seen = ws_master.cell(row=rownum, column=col_idx["last_seen_et"]).value
        if isinstance(last_seen, datetime) and last_seen < stale_cutoff:
            ws_master.cell(row=rownum, column=col_idx["status"]).value = "Stale"

            # Copy to archive
            arc_values = [ws_master.cell(row=rownum, column=i+1).value for i in range(len(MASTER_HDR))]
            ws_archive.append(arc_values)
            to_archive.append(rownum)

            stale_row = {header[i]: arc_values[i] for i in range(len(header))}
            actions.append(Action("Stale", stale_row))

    # Delete archived rows
    for rownum in sorted(to_archive, reverse=True):
        ws_master.delete_rows(rownum, 1)

    # Deduplicate archive
    deduplicate_archive(wb)

    # Apply styling
    styler = ExcelStyler()
    styler.apply_header_style(ws_master)
    apply_table_style(ws_master)
    apply_conditional_formatting(ws_master)
    auto_size_columns(ws_master)

    # Log actions
    for a in actions:
        row_for_log = a.row.copy()
        for k in ["first_seen_et", "last_seen_et", "publish_dt_et", "due_dt_et"]:
            row_for_log[k] = to_excel_naive(row_for_log.get(k))
        ws_log.append([ts_run_xl, a.action] + [row_for_log.get(k) for k in MASTER_HDR])

    # Update analytics
    if ws_analytics:
        counts = {
            "New": sum(1 for a in actions if a.action == "New"),
            "Updated": sum(1 for a in actions if a.action == "Updated"),
            "Unchanged": sum(1 for a in actions if a.action == "Unchanged"),
            "Stale": sum(1 for a in actions if a.action == "Stale"),
        }
        ws_analytics.append([
            ts_run_xl,
            counts["New"],
            counts["Updated"],
            counts["Unchanged"],
            counts["Stale"],
            sum(counts.values())
        ])

    # Create backup before saving
    create_workbook_backup(wb_path)

    # Save with error handling
    try:
        wb.save(wb_path)
        logger.info(f"Saved workbook: {wb_path}")
        logger.info(f"Actions - New: {counts['New']}, Updated: {counts['Updated']}, "
                   f"Unchanged: {counts['Unchanged']}, Stale: {counts['Stale']}")
    except PermissionError:
        logger.error("Cannot save workbook - file is open. Please close it and try again.")
        raise
    except Exception as exc:
        logger.error(f"Failed to save workbook: {exc}")
        raise

def apply_table_style(ws):
    """Apply Excel table style."""
    if ws.max_row < 1 or ws.max_column < 1:
        return

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Check for existing table
    existing_tables = getattr(ws, "tables", {})
    tbl = existing_tables.get("tbl_opps") if existing_tables else None

    if tbl:
        tbl.ref = ref
    else:
        tbl = Table(displayName="tbl_opps", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

    ws.freeze_panes = "A2"

def apply_conditional_formatting(ws):
    """Apply status-based conditional formatting."""
    header = [c.value for c in ws[1]]
    if "status" not in header:
        return

    status_col = header.index("status") + 1
    col_letter = get_column_letter(status_col)
    nrows = ws.max_row

    if nrows < 2:
        return

    # Status colors
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    grey = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Add rules
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{nrows}",
        FormulaRule(formula=[f'${col_letter}2="New"'], stopIfTrue=False, fill=green)
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{nrows}",
        FormulaRule(formula=[f'${col_letter}2="Updated"'], stopIfTrue=False, fill=amber)
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{nrows}",
        FormulaRule(formula=[f'${col_letter}2="Stale"'], stopIfTrue=False, fill=grey)
    )

def auto_size_columns(ws, max_width=60):
    """Auto-size columns based on content."""
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_length = 0

        for row in range(1, min(ws.max_row + 1, 100)):  # Sample first 100 rows
            cell_value = str(ws.cell(row=row, column=col).value or "")
            max_length = max(max_length, len(cell_value))

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[col_letter].width = adjusted_width

# -------------------- CLI Commands --------------------
def run_update(days_ago: int, skip_details: bool = False):
    """Run standard update process."""
    logger.info(f"Starting update for {days_ago} days ago")

    staging = emma_scrape_enhanced(
        days_ago=days_ago,
        max_pages=MAX_PAGES,
        sleep_s=SLEEP_BETWEEN,
        fetch_details=not skip_details,
        adaptive_rate=True
    )

    merge_into_excel_enhanced(staging, WORKBOOK_PATH)

    logger.info(f"Update complete: {WORKBOOK_PATH} (target: {days_ago} days ago)")

def handle_scrape_command(args):
    """Handle scrape subcommand."""
    if hasattr(args, 'multi_date') and args.multi_date:
        days_list = [int(d.strip()) for d in args.multi_date.split(',')]
        logger.info(f"Multi-date scrape: {days_list}")

        for day in days_list:
            logger.info(f"Scraping day {day}...")
            run_update(day, args.skip_details)
    else:
        days_ago = args.days_ago if hasattr(args, 'days_ago') else DAYS_AGO
        run_update(days_ago, args.skip_details)

def handle_archive_command(args):
    """Handle archive subcommand."""
    logger.info(f"Manual archive: records older than {args.days} days")

    wb = load_workbook_safe(WORKBOOK_PATH)

    # Archive logic here (simplified)
    ws_master = wb["Master"]
    ws_archive = wb["Archive"]

    cutoff = now_et() - timedelta(days=args.days)
    cutoff_xl = to_excel_naive(cutoff)

    header = [c.value for c in ws_master[1]]
    last_seen_col = header.index("last_seen_et") + 1

    to_delete = []
    for row_num in range(2, ws_master.max_row + 1):
        last_seen = ws_master.cell(row=row_num, column=last_seen_col).value
        if isinstance(last_seen, datetime) and last_seen < cutoff_xl:
            # Copy to archive
            row_data = [ws_master.cell(row=row_num, column=i+1).value for i in range(len(header))]
            ws_archive.append(row_data)
            to_delete.append(row_num)

    # Delete from master
    for row_num in sorted(to_delete, reverse=True):
        ws_master.delete_rows(row_num, 1)

    # Deduplicate archive
    deduplicate_archive(wb)

    wb.save(WORKBOOK_PATH)
    logger.info(f"Archived {len(to_delete)} records")

def handle_report_command(args):
    """Handle report subcommand."""
    logger.info(f"Generating {args.format} report: {args.output}")

    wb = load_workbook(WORKBOOK_PATH, read_only=True)
    generate_analytics_report(wb, args.output, args.format)

# -------------------- Main Entry Point --------------------
def main():
    """Enhanced CLI with all features."""
    parser = argparse.ArgumentParser(
        prog="emma_scraper",
        description="Enhanced eMMA Excel Updater",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Environment variables:
  EMMA_XLSX        Path to Excel workbook
  DAYS_AGO         Days prior to scrape (default: 0)
  STALE_AFTER_D    Archive after N days (default: 7)
  MAX_PAGES        Max pages to scrape (default: 50)
  SLEEP_BETWEEN    Request delay (default: 1.0)
  LOG_LEVEL        Log level (default: INFO)

Examples:
  %(prog)s                              # Run with defaults
  %(prog)s --days-ago 1                 # Yesterday's listings
  %(prog)s --json-logs                  # JSON logging
  %(prog)s scrape --multi-date 0,1,2    # Multi-day scrape
  %(prog)s archive --days 30            # Archive old records
  %(prog)s report --format json         # Generate analytics
        """
    )

    # Global arguments
    parser.add_argument("--days-ago", type=int, default=None,
                       help="Days prior to today")
    parser.add_argument("--skip-details", action="store_true",
                       help="Skip detail page fetching")
    parser.add_argument("--log-level", type=str, default=None,
                       choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                       help="Log level")
    parser.add_argument("--json-logs", action="store_true",
                       help="Use JSON logging")
    parser.add_argument("--workbook", type=str, default=None,
                       help="Path to workbook")

    # Subcommands
    subparsers = parser.add_subparsers(dest='command', help='Commands')

    # Scrape command
    scrape_parser = subparsers.add_parser('scrape', help='Scrape listings')
    scrape_parser.add_argument("--multi-date", type=str,
                              help="Comma-separated days (e.g., '0,1,2')")
    scrape_parser.add_argument("--days-ago", type=int, default=None)
    scrape_parser.add_argument("--skip-details", action="store_true")

    # Archive command
    archive_parser = subparsers.add_parser('archive', help='Archive records')
    archive_parser.add_argument("--days", type=int, default=STALE_AFTER_D,
                               help="Archive older than N days")

    # Report command
    report_parser = subparsers.add_parser('report', help='Generate report')
    report_parser.add_argument("--output", type=str, default="analytics.csv",
                              help="Output file")
    report_parser.add_argument("--format", choices=["csv", "xlsx", "json"],
                              default="csv", help="Output format")

    args = parser.parse_args()

    # Configure logging
    log_level = args.log_level or LOG_LEVEL
    configure_logging(log_level, args.json_logs)

    # Override globals if specified
    global WORKBOOK_PATH, DAYS_AGO
    if args.workbook:
        WORKBOOK_PATH = args.workbook
    if hasattr(args, 'days_ago') and args.days_ago is not None:
        DAYS_AGO = args.days_ago

    # Execute command
    try:
        if args.command == 'scrape':
            handle_scrape_command(args)
        elif args.command == 'archive':
            handle_archive_command(args)
        elif args.command == 'report':
            handle_report_command(args)
        else:
            # Default behavior
            days = args.days_ago if args.days_ago is not None else DAYS_AGO
            run_update(days, args.skip_details)

    except KeyboardInterrupt:
        logger.info("Interrupted by user")
    except Exception as exc:
        logger.error(f"Error: {exc}", exc_info=True)
        raise

if __name__ == "__main__":
    main()