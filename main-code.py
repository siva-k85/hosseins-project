"""
eMMA -> Excel updater




- Scrapes Maryland eMMA public listings using requests+bs4
- Merges into one Excel workbook:
  * Master: current rows (<= 7 days old), styled table
  * Log: append-only history of actions
  * Archive: pruned (stale) rows
  * Refs: optional rules you can fill manually
- No Chrome/driver, no Selenium




Env vars you can set:
- EMMA_XLSX (default: opportunities.xlsx)
- DAYS_AGO (default: 2)       # 1=yesterday, 2=day before, etc.
- STALE_AFTER_D (default: 7)  # days after which untouched rows are archived
- MAX_PAGES (default: 50)
- SLEEP_BETWEEN (default: 1.0)
- LOG_LEVEL (default: INFO)
- TIMEOUT_SECONDS (default: 30)
- USER_AGENT (default provided)
"""




import argparse
import os
import re
import time
import logging
from hashlib import blake2b
from dataclasses import dataclass
from datetime import datetime, timedelta
from urllib.parse import urljoin
from zipfile import BadZipFile




import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from typing import Optional, Tuple


# -------------------- Config (env-overridable) --------------------

def validate_parameters():
    """Validate configuration parameters with helpful error messages."""
    errors = []

    # Validate DAYS_AGO
    days_ago = os.getenv("DAYS_AGO", "0")
    try:
        days_ago_int = int(days_ago)
        if days_ago_int < 0:
            errors.append(f"DAYS_AGO must be non-negative, got {days_ago_int}")
    except ValueError:
        errors.append(f"DAYS_AGO must be an integer, got '{days_ago}'")

    # Validate STALE_AFTER_D
    stale_after = os.getenv("STALE_AFTER_D", "7")
    try:
        stale_after_int = int(stale_after)
        if stale_after_int <= 0:
            errors.append(f"STALE_AFTER_D must be positive, got {stale_after_int}")
    except ValueError:
        errors.append(f"STALE_AFTER_D must be an integer, got '{stale_after}'")

    # Validate MAX_PAGES
    max_pages = os.getenv("MAX_PAGES", "50")
    try:
        max_pages_int = int(max_pages)
        if max_pages_int <= 0:
            errors.append(f"MAX_PAGES must be positive, got {max_pages_int}")
    except ValueError:
        errors.append(f"MAX_PAGES must be an integer, got '{max_pages}'")

    # Validate SLEEP_BETWEEN
    sleep_between = os.getenv("SLEEP_BETWEEN", "1.0")
    try:
        sleep_between_float = float(sleep_between)
        if sleep_between_float < 0:
            errors.append(f"SLEEP_BETWEEN must be non-negative, got {sleep_between_float}")
    except ValueError:
        errors.append(f"SLEEP_BETWEEN must be a number, got '{sleep_between}'")

    # Validate TIMEOUT_SECONDS
    timeout = os.getenv("TIMEOUT_SECONDS", "30")
    try:
        timeout_int = int(timeout)
        if timeout_int <= 0:
            errors.append(f"TIMEOUT_SECONDS must be positive, got {timeout_int}")
    except ValueError:
        errors.append(f"TIMEOUT_SECONDS must be an integer, got '{timeout}'")

    # Validate LOG_LEVEL
    log_level = os.getenv("LOG_LEVEL", "INFO").upper()
    valid_levels = ["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"]
    if log_level not in valid_levels:
        errors.append(f"LOG_LEVEL must be one of {valid_levels}, got '{log_level}'")

    if errors:
        raise ValueError("Configuration validation failed:\n" + "\n".join(f"  - {e}" for e in errors))

# Cross-platform path handling
def get_default_workbook_path():
    """Get platform-appropriate default workbook path."""
    if os.name == 'nt':  # Windows
        return r"C:\Users\hkhoshhal001\Guidehouse\New York Mid-Atlantic (NYMA) - 05. Scanning resources\Automated Scanning\opportunities.xlsx"
    else:  # Unix/Linux/Mac
        home = os.path.expanduser("~")
        return os.path.join(home, "Documents", "emma_opportunities.xlsx")

WORKBOOK_PATH = os.getenv("EMMA_XLSX", get_default_workbook_path())

# Validate parameters before using them
validate_parameters()

DAYS_AGO        = int(os.getenv("DAYS_AGO", "0"))     # 1=yesterday, 2=day before, etc.
STALE_AFTER_D   = int(os.getenv("STALE_AFTER_D", "7"))
MAX_PAGES       = int(os.getenv("MAX_PAGES", "50"))
SLEEP_BETWEEN   = float(os.getenv("SLEEP_BETWEEN", "1.0"))
LOG_LEVEL       = os.getenv("LOG_LEVEL", "INFO").upper()
TIMEOUT_SECONDS = int(os.getenv("TIMEOUT_SECONDS", "30"))
USER_AGENT      = os.getenv("USER_AGENT", "Mozilla/5.0 (compatible; MD-EmmaScraper/1.0)")




BASE = "https://emma.maryland.gov"
BROWSE_URL = f"{BASE}/page.aspx/en/rfp/request_browse_public"

# -------------------- Timezone (ET) --------------------
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    ET_TZ = ZoneInfo("America/New_York")
except Exception:
    ET_TZ = None




def now_et():
    return datetime.now(ET_TZ) if ET_TZ else datetime.now()




def localize_et(dt: datetime) -> datetime:
    return dt.replace(tzinfo=ET_TZ) if ET_TZ else dt




def to_excel_naive(dt):
    """Excel/openpyxl cannot store tz-aware datetimes. Return naive (no tzinfo)."""
    if dt is None:
        return None
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt




# -------------------- Logging --------------------
# Create a named logger (configured via configure_logging)
logger = logging.getLogger("emma_scraper")


def configure_logging(level_name: str):
    level = getattr(logging, level_name.upper(), None)
    if not isinstance(level, int):
        logger.warning("Unknown LOG_LEVEL '%s'; defaulting to INFO", level_name)
        level = logging.INFO

    root = logging.getLogger()
    if not root.handlers:
        handler = logging.StreamHandler()
        handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        root.addHandler(handler)
    root.setLevel(level)
    logger.setLevel(level)
    logging.captureWarnings(True)


configure_logging(LOG_LEVEL)




# -------------------- HTTP session --------------------
def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT,
                      "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"})
    retries = Retry(total=5, backoff_factor=0.6,
                    status_forcelist=[429,500,502,503,504],
                    allowed_methods=["GET","POST"], raise_on_status=False)
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.mount("http://", HTTPAdapter(max_retries=retries))
    return s

# -------------------- eMMA scraping --------------------
TS_PATTERN = re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s+(AM|PM)\b")

DETAIL_SUMMARY_LABELS = ["summary", "description", "project description", "scope", "overview"]
DETAIL_OFFICER_LABELS = ["procurement officer", "buyer", "contact", "procurement contact", "issuing officer"]
DETAIL_EMAIL_LABELS = ["email", "e-mail", "contact email", "buyer email"]
DETAIL_INSTRUCTIONS_LABELS = ["instruction", "guideline", "submission", "note", "requirement", "special instruction"]
DETAIL_GOALS_LABELS = ["goal", "participation", "mbe", "dbe", "sbe", "wbe", "small business", "program goal"]
DETAIL_DUE_LABELS = ["due date", "bid due date", "proposal due date", "response due date", "closing date", "closing time"]
DETAIL_TIMESTAMP_FORMATS = [
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y",
    "%Y-%m-%d %H:%M:%S",
]

DETAIL_FIELD_KEYS = [
    "solicitation_summary",
    "procurement_officer_buyer",
    "contact_email",
    "additional_instructions",
    "procurement_program_goals",
]

EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")

PUBLISH_DT_FORMATS = [
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%Y-%m-%d %H:%M:%S",
    "%m/%d/%y %I:%M %p",
]

LISTING_HEADER_ALIASES = {
    "title": {
        "title", "solicitation", "solicitation title", "project title", "opportunity",
    },
    "solicitation_id": {
        "solicitation id", "solicitation #", "solicitation number", "reference #", "document number",
        "document id", "document #", "rfp number", "bid number", "contract number",
    },
    "category": {
        "category", "procurement category", "commodity", "product category",
    },
    "procurement_method": {
        "procurement method", "method", "contract type", "procurement type",
    },
    "agency": {
        "agency", "issuing agency", "department", "agency/department",
    },
    "publish_dt": {
        "publish date", "posting date", "posted", "issue date", "posting date/time",
    },
    "due_dt": {
        "due date", "bid due date", "proposal due date", "response due date", "closing date", "closing time",
    },
}

DEFAULT_COLUMN_INDEXES = {
    "title": 2,
    "category": 6,
    "procurement_method": 7,
    "agency": 8,
    "publish_dt": 5,
    "solicitation_id": 1,
    "due_dt": 10,
}




def parse_hidden_fields(soup: BeautifulSoup) -> dict:
    fields = {}
    for name in ["__VIEWSTATE","__EVENTVALIDATION","__VIEWSTATEGENERATOR","__EVENTTARGET","__EVENTARGUMENT"]:
        el = soup.find("input", {"name": name})
        if el and el.has_attr("value"):
            fields[name] = el["value"]
    return fields


def _empty_detail_fields() -> dict:
    return {key: "" for key in DETAIL_FIELD_KEYS}


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _iter_labeled_values(soup: BeautifulSoup):
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            label = _normalize_text(cells[0].get_text(" ", strip=True))
            value = _normalize_text(cells[1].get_text(" ", strip=True))
            if label and value:
                yield label.lower(), value

    for dt_tag in soup.find_all("dt"):
        label = _normalize_text(dt_tag.get_text(" ", strip=True))
        dd = dt_tag.find_next_sibling("dd")
        if dd:
            value = _normalize_text(dd.get_text(" ", strip=True))
            if label and value:
                yield label.lower(), value


def _build_column_map(header_cells) -> dict:
    mapping = {}
    for idx, cell in enumerate(header_cells or []):
        label = _normalize_text(cell.get_text(" ", strip=True)).lower()
        if not label:
            continue
        for canonical, aliases in LISTING_HEADER_ALIASES.items():
            if label in aliases:
                mapping[canonical] = idx
                break
    return mapping


def _extract_value_by_labels(soup: BeautifulSoup, labels: list[str]) -> str:
    lowered = [label.lower() for label in labels]
    for label, value in _iter_labeled_values(soup):
        if any(keyword in label for keyword in lowered):
            return value
    return ""


def _first_paragraph_text(soup: BeautifulSoup) -> str:
    for p in soup.find_all("p"):
        text = _normalize_text(p.get_text(" ", strip=True))
        if len(text) >= 20:
            return text
    return ""


def _extract_solicitation_summary(soup: BeautifulSoup) -> str:
    value = _extract_value_by_labels(soup, DETAIL_SUMMARY_LABELS)
    if value:
        return value
    return _first_paragraph_text(soup)


def _extract_procurement_officer(soup: BeautifulSoup) -> str:
    return _extract_value_by_labels(soup, DETAIL_OFFICER_LABELS)


def _extract_contact_email(soup: BeautifulSoup) -> str:
    value = _extract_value_by_labels(soup, DETAIL_EMAIL_LABELS)
    if value:
        match = EMAIL_REGEX.search(value)
        if match:
            return match.group(0)
        return value
    match = EMAIL_REGEX.search(soup.get_text(" ", strip=True))
    return match.group(0) if match else ""


def _extract_additional_instructions(soup: BeautifulSoup) -> str:
    value = _extract_value_by_labels(soup, DETAIL_INSTRUCTIONS_LABELS)
    if value:
        return value
    matches = []
    lowered_keywords = [k.lower() for k in DETAIL_INSTRUCTIONS_LABELS]
    for element in soup.find_all(["p", "li"]):
        text = _normalize_text(element.get_text(" ", strip=True))
        if not text:
            continue
        lower = text.lower()
        if any(keyword in lower for keyword in lowered_keywords):
            matches.append(text)
    return " ".join(matches)


def _extract_program_goals(soup: BeautifulSoup) -> str:
    value = _extract_value_by_labels(soup, DETAIL_GOALS_LABELS)
    if value:
        return value
    matches = []
    lowered_keywords = [k.lower() for k in DETAIL_GOALS_LABELS]
    for element in soup.find_all(["p", "li"]):
        text = _normalize_text(element.get_text(" ", strip=True))
        if not text:
            continue
        lower = text.lower()
        if any(keyword in lower for keyword in lowered_keywords):
            matches.append(text)
    return " ".join(matches)


def _extract_due_datetime(soup: BeautifulSoup) -> tuple[str, Optional[datetime]]:
    value = _extract_value_by_labels(soup, DETAIL_DUE_LABELS)
    if not value:
        text = soup.get_text(" ", strip=True)
        match = TS_PATTERN.search(text)
        if match:
            value = match.group(0)
    value = _normalize_text(value)
    if not value:
        return "", None
    for fmt in DETAIL_TIMESTAMP_FORMATS:
        try:
            dt = datetime.strptime(value, fmt)
            return value, localize_et(dt)
        except Exception:
            continue
    return value, None


def scrape_detail_page(session: requests.Session, url: str) -> dict:
    detail_payload = _empty_detail_fields()
    detail_payload.update({
        "detail_due_text": "",
        "due_dt_et": None,
        "__fetched__": False,
    })

    if not url:
        return detail_payload

    try:
        response = session.get(url, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
    except Exception as exc:
        logger.warning("Failed to fetch detail page %s: %s", url, exc)
        return detail_payload

    soup = BeautifulSoup(response.text, "html.parser")
    detail_payload["solicitation_summary"] = _extract_solicitation_summary(soup)
    detail_payload["procurement_officer_buyer"] = _extract_procurement_officer(soup)
    detail_payload["contact_email"] = _extract_contact_email(soup)
    detail_payload["additional_instructions"] = _extract_additional_instructions(soup)
    detail_payload["procurement_program_goals"] = _extract_program_goals(soup)
    detail_payload["detail_due_text"], detail_payload["due_dt_et"] = _extract_due_datetime(soup)
    detail_payload["__fetched__"] = True
    return detail_payload


def _make_record_id(row: dict) -> str:
    solicitation_id = (row.get("solicitation_id") or "").strip()
    if solicitation_id:
        return solicitation_id
    url = row.get("url") or ""
    if url:
        match = re.search(r"/extranet/(\d+)", url)
        if match:
            return match.group(1)
    seed = (url + (row.get("title") or "")).encode()
    return f"emma_{blake2b(seed, digest_size=8).hexdigest()}"


def _deduplicate_rows(rows: list[dict]):
    deduped = []
    seen = set()
    duplicates = 0
    for row in rows:
        key = (
            (row.get("title") or "").strip().lower(),
            (row.get("agency") or "").strip().lower(),
            row.get("_publish_dt_key"),
            (row.get("solicitation_id") or "").strip().lower(),
        )
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        deduped.append(row)
    return deduped, duplicates


def extract_rows(soup: BeautifulSoup) -> list[dict]:
    table = soup.select_one("table.iv-grid-view")
    if not table:
        for t in soup.find_all("table"):
            classes = " ".join(t.get("class", []))
            if "iv-grid" in classes or "iv-grid-view" in classes or "very compact" in classes:
                table = t
                break
    if not table:
        return []




    rows = []
    tbody = table.find("tbody")
    body_rows = tbody.find_all("tr") if tbody else table.find_all("tr")

    header_row = None
    header_cells = []
    thead = table.find("thead")
    if thead:
        header_row = thead.find("tr")
    elif body_rows:
        first = body_rows[0]
        if first.find_all("th"):
            header_row = first
    if header_row:
        header_cells = header_row.find_all(["th", "td"])
        body_rows = [row for row in body_rows if row is not header_row]

    column_map = _build_column_map(header_cells)
    fallback_keys: set[str] = set()

    def idx_for(key: str):
        if key in column_map:
            return column_map[key]
        if key in DEFAULT_COLUMN_INDEXES:
            fallback_keys.add(key)
            return DEFAULT_COLUMN_INDEXES[key]
        return None

    for tr in body_rows:
        tds = tr.find_all("td")
        if not tds:
            continue

        title_idx = idx_for("title") or 0
        if title_idx >= len(tds):
            continue
        title_cell = tds[title_idx]
        a = title_cell.find("a")
        title = a.get_text(strip=True) if a else title_cell.get_text(strip=True)
        link = urljoin(BASE, a["href"]) if (a and a.has_attr("href")) else None

        def value_from_idx(key: str, default: str = ""):
            idx = idx_for(key)
            if idx is None or idx >= len(tds):
                return default
            return tds[idx].get_text(" ", strip=True)

        category = value_from_idx("category")
        method = value_from_idx("procurement_method")
        agency = value_from_idx("agency")
        solicitation_id = value_from_idx("solicitation_id")

        publish_text = value_from_idx("publish_dt")
        if not publish_text:
            for td in tds:
                txt = td.get_text(" ", strip=True)
                m = TS_PATTERN.search(txt)
                if m:
                    publish_text = m.group(0)
                    break

        due_text = value_from_idx("due_dt")

        rows.append({
            "title": title or "",
            "url": link,
            "category": category,
            "procurement_method": method,
            "agency": agency,
            "publish_dt_raw": publish_text or "",
            "due_dt_raw": due_text or "",
            "solicitation_id": solicitation_id or "",
        })

    if fallback_keys:
        logger.warning(
            "extract_rows falling back to positional indices for columns: %s",
            ", ".join(sorted(fallback_keys))
        )
    return rows




def find_next_postback(soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str]]:
    for a in soup.find_all("a", href=True):
        m = re.search(r"__doPostBack\('([^']*)','([^']*)'\)", a["href"])
        if m and (a.get_text(strip=True) or "").lower() in {"next",">",">>"}:
            return m.group(1), m.group(2)
    # Fallback to highest page number
    candidates = []
    for a in soup.find_all("a", href=True):
        m = re.search(r"__doPostBack\('([^']*)','([^']*)'\)", a["href"])
        label = a.get_text(strip=True)
        if m and label.isdigit():
            candidates.append((int(label), m.group(1), m.group(2)))
    if candidates:
        _, et, ea = sorted(candidates, key=lambda x:x[0])[-1]
        return et, ea
    return None, None




def parse_publish_dt(raw: str) -> Optional[datetime]:
    if not raw:
        return None
    cleaned = raw.strip()
    for fmt in PUBLISH_DT_FORMATS:
        try:
            dt = datetime.strptime(cleaned, fmt)
            logger.debug("parse_publish_dt matched format '%s' for value '%s'", fmt, cleaned)
            return localize_et(dt)
        except ValueError:
            continue
    logger.debug("parse_publish_dt failed for value '%s'", cleaned)
    return None




def emma_scrape(DAYS_AGO: int, max_pages:int=50, sleep_s:float=1.0, fetch_details: bool = True) -> list[dict]:
    """Return a list of normalized records for the target ET date."""
    ses = make_session()
    r = ses.get(BROWSE_URL, timeout=TIMEOUT_SECONDS)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    all_rows = []
    pages = 0
    seen = set()

    def page_fp(s):  # guard against repeats
        t = (s.select_one("table.iv-grid-view") or s.find("table"))
        text = (t.get_text(" ", strip=True) if t else s.get_text(" ", strip=True))[:20000]
        return blake2b(text.encode(), digest_size=16).hexdigest()

    while True:
        fp = page_fp(soup)
        if fp in seen:
            break
        seen.add(fp)

        rows = extract_rows(soup)
        all_rows.extend(rows)
        pages += 1
        if pages >= max_pages:
            break

        et, ea = find_next_postback(soup)
        if not et:
            break

        fields = parse_hidden_fields(soup)
        fields["__EVENTTARGET"]  = et
        fields["__EVENTARGUMENT"] = ea
        time.sleep(sleep_s)
        r = ses.post(BROWSE_URL, data=fields, timeout=TIMEOUT_SECONDS)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

    if fetch_details and all_rows:
        success = 0
        failure = 0
        total = len(all_rows)
        for idx, row in enumerate(all_rows, start=1):
            detail = scrape_detail_page(ses, row.get("url"))
            if detail.get("__fetched__"):
                success += 1
            else:
                failure += 1
            for key in DETAIL_FIELD_KEYS:
                row[key] = detail.get(key, "")
            if detail.get("due_dt_et") and not row.get("due_dt_et"):
                row["due_dt_et"] = detail["due_dt_et"]
            if idx % 10 == 0 or idx == total:
                logger.info("Fetched %d/%d detail pages...", idx, total)
            if idx < total:
                time.sleep(sleep_s)
        logger.info("Detail page fetch summary: success=%d failure=%d", success, failure)
    else:
        for row in all_rows:
            for key in DETAIL_FIELD_KEYS:
                row.setdefault(key, "")

    for row in all_rows:
        publish_raw = row.pop("publish_dt_raw", "")
        publish_dt = parse_publish_dt(publish_raw)
        row["publish_dt_et"] = publish_dt
        row["_publish_dt_key"] = publish_dt.isoformat() if publish_dt else publish_raw.strip().lower()
        row["source"] = "emma"

        due_raw = row.pop("due_dt_raw", "")
        if not row.get("due_dt_et"):
            row["due_dt_et"] = parse_publish_dt(due_raw)

        row["solicitation_id"] = (row.get("solicitation_id") or "").strip()
        row.setdefault("due_dt_et", None)

        row.setdefault("tags", "")
        row.setdefault("score_bd_fit", "")

    all_rows, duplicate_count = _deduplicate_rows(all_rows)
    if duplicate_count:
        logger.info("Filtered %d duplicate rows prior to staging.", duplicate_count)

    for row in all_rows:
        row["record_id"] = _make_record_id(row)
        row.pop("_publish_dt_key", None)

    target_date = (now_et().date() - timedelta(days=DAYS_AGO))
    staging = [r for r in all_rows if r.get("publish_dt_et") and r["publish_dt_et"].date() == target_date]
    return staging







# -------------------- Excel helpers --------------------
MASTER_HDR = [
    "source","record_id","url","first_seen_et","last_seen_et",
    "title","agency","category","procurement_method","publish_dt_et","due_dt_et","solicitation_id",
    "solicitation_summary","procurement_officer_buyer","contact_email","additional_instructions","procurement_program_goals",
    "status","tags","score_bd_fit"
]


def create_workbook_backup(original_path: str, max_backups: int = 5) -> str:
    """
    Create a timestamped backup of the workbook file.

    Args:
        original_path: Path to the original workbook file
        max_backups: Maximum number of backups to keep (default: 5)

    Returns:
        Path to the created backup file
    """
    if not os.path.exists(original_path):
        logger.debug("Original workbook '%s' doesn't exist, skipping backup", original_path)
        return ""

    try:
        # Create backup directory
        backup_dir = os.path.join(os.path.dirname(original_path), "backups")
        os.makedirs(backup_dir, exist_ok=True)

        # Generate timestamped backup filename
        base_name = os.path.splitext(os.path.basename(original_path))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{base_name}_backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)

        # Copy the file
        import shutil
        shutil.copy2(original_path, backup_path)

        logger.info("Created workbook backup: %s", backup_path)

        # Cleanup old backups
        cleanup_old_backups(backup_dir, base_name, max_backups)

        return backup_path

    except (OSError, IOError) as exc:
        logger.warning("Failed to create backup of '%s': %s", original_path, exc)
        return ""


def cleanup_old_backups(backup_dir: str, base_name: str, max_backups: int) -> None:
    """
    Remove old backup files, keeping only the most recent ones.

    Args:
        backup_dir: Directory containing backup files
        base_name: Base name of the workbook (without extension)
        max_backups: Maximum number of backups to keep
    """
    try:
        if not os.path.exists(backup_dir):
            return

        # Find all backup files for this workbook
        backup_files = []

        for filename in os.listdir(backup_dir):
            if filename.startswith(f"{base_name}_backup_") and filename.endswith(".xlsx"):
                filepath = os.path.join(backup_dir, filename)
                if os.path.isfile(filepath):
                    backup_files.append((filepath, os.path.getmtime(filepath)))

        # Sort by modification time (newest first)
        backup_files.sort(key=lambda x: x[1], reverse=True)

        # Remove old backups beyond max_backups limit
        files_to_remove = backup_files[max_backups:]
        removed_count = 0

        for filepath, _ in files_to_remove:
            try:
                os.remove(filepath)
                removed_count += 1
                logger.debug("Removed old backup: %s", filepath)
            except OSError as exc:
                logger.warning("Failed to remove old backup '%s': %s", filepath, exc)

        if removed_count > 0:
            logger.info("Cleaned up %d old backup files in %s", removed_count, backup_dir)

    except OSError as exc:
        logger.warning("Failed to cleanup old backups in '%s': %s", backup_dir, exc)




def init_workbook_if_needed(path:str):
    if os.path.exists(path):
        return
    wb = Workbook()
    # create sheets in order
    ws_master = wb.active
    ws_master.title = "Master"
    ws_master.append(MASTER_HDR)




    ws_log = wb.create_sheet("Log")
    ws_log.append(["run_ts_et","action"] + MASTER_HDR)




    wb.create_sheet("Refs")
    wb.create_sheet("Archive")




    create_workbook_backup(path)
    wb.save(path)




def load_wb(path:str):
    # Create if missing
    init_workbook_if_needed(path)
    # Try opening; if corrupt, rebuild
    try:
        return load_workbook(path)
    except (BadZipFile, InvalidFileException):
        logger.warning("File at %s was not a valid Excel. Reinitializing.", path)
        try:
            os.remove(path)
        except Exception:
            pass
        init_workbook_if_needed(path)
        return load_workbook(path)
    except (PermissionError, OSError) as exc:
        raise RuntimeError(f"Cannot open workbook '{path}': {exc}") from exc


def ensure_header(ws, expected_header):
    current_header = [c.value for c in ws[1]] if ws.max_row else []
    if current_header == expected_header:
        return

    preserved_rows = []
    if current_header:
        col_map = {name: idx for idx, name in enumerate(current_header) if name}
        for row_idx in range(2, ws.max_row + 1):
            data = {}
            for name, idx in col_map.items():
                data[name] = ws.cell(row=row_idx, column=idx + 1).value
            preserved_rows.append(data)

    ws.delete_rows(1, ws.max_row)
    ws.append(expected_header)
    for data in preserved_rows:
        ws.append([data.get(col) for col in expected_header])




def ws_to_index(ws, key_col_name="record_id") -> dict:
    # build index of existing Master rows: record_id -> row number
    header = [c.value for c in ws[1]]
    col_idx = {name: header.index(name)+1 for name in header}
    idx = {}
    for row in range(2, ws.max_row+1):
        rid = ws.cell(row=row, column=col_idx[key_col_name]).value
        if rid:
            idx[rid] = row
    return header, col_idx, idx




def ensure_master_table_style(ws):
    """
    Ensure there is a single styled Excel Table named tbl_opps that
    spans all data in ws. If a table already exists, resize it.
    """
    last_row = ws.max_row
    last_col = ws.max_column
    if last_row < 1 or last_col < 1:
        return




    ref = f"A1:{get_column_letter(last_col)}{last_row}"




    # openpyxl exposes tables via ws.tables (dict-like). If present, resize; if not, create.
    existing_tables = getattr(ws, "tables", {})
    tbl = existing_tables.get("tbl_opps") if existing_tables else None




    if tbl:
        tbl.ref = ref
    else:
        tbl = Table(displayName="tbl_opps", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)




    # Freeze header row for readability
    ws.freeze_panes = "A2"




def auto_col_widths(ws, max_width=60):
    # auto size columns based on content (simple heuristic)
    for col in range(1, ws.max_column+1):
        values = [str(ws.cell(row=r, column=col).value or "") for r in range(1, ws.max_row+1)]
        width = min(max((len(v) for v in values), default=10) + 2, max_width)
        ws.column_dimensions[get_column_letter(col)].width = width




def apply_status_conditional_formats(ws):
    # Simple conditional fill by status (assumes status is in column 'L' per MASTER_HDR)
    header = [c.value for c in ws[1]]
    if "status" not in header: return
    c = header.index("status") + 1
    col_letter = get_column_letter(c)
    nrows = ws.max_row
    if nrows < 2: return




    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    grey  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")




    # New
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{nrows}",
        FormulaRule(formula=[f'${col_letter}2="New"'], stopIfTrue=False, fill=green))
    # Updated
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{nrows}",
        FormulaRule(formula=[f'${col_letter}2="Updated"'], stopIfTrue=False, fill=amber))
    # Stale
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{nrows}",
        FormulaRule(formula=[f'${col_letter}2="Stale"'], stopIfTrue=False, fill=grey))




def row_dict(ws, row_num:int) -> dict:
    header = [c.value for c in ws[1]]
    return {name: ws.cell(row=row_num, column=i+1).value for i, name in enumerate(header)}




def rows_equal(existing:dict, incoming:dict) -> bool:
    # compare user-visible business fields only
    keys = [
        "title","agency","category","procurement_method","publish_dt_et","due_dt_et","url",
        "solicitation_id","solicitation_summary","procurement_officer_buyer",
        "contact_email","additional_instructions","procurement_program_goals"
    ]
    for k in keys:
        if existing.get(k) != incoming.get(k):
            return False
    return True




# -------------------- Merge pipeline --------------------
@dataclass
class Action:
    action: str
    row: dict




def merge_into_excel(staging: list[dict]):
    ts_run = now_et()
    ts_run_xl = to_excel_naive(ts_run)




    wb = load_wb(WORKBOOK_PATH)
    ws_master = wb["Master"]
    ws_log = wb["Log"]
    ws_archive = wb["Archive"]




    ensure_header(ws_master, MASTER_HDR)
    ensure_header(ws_archive, MASTER_HDR)
    ensure_header(ws_log, ["run_ts_et","action"] + MASTER_HDR)




    header, col_idx, index = ws_to_index(ws_master)




    actions: list[Action] = []
    touched_ids = set()




    # Convert staging rows to Master schema
    for r in staging:
        row = {
            "source": "emma",
            "record_id": r["record_id"],
            "url": r["url"],
            "first_seen_et": None,  # set on insert
            "last_seen_et": ts_run_xl,
            "title": r["title"],
            "agency": r["agency"],
            "category": r["category"],
            "procurement_method": r["procurement_method"],
            "publish_dt_et": r["publish_dt_et"],
            "due_dt_et": r.get("due_dt_et"),
            "solicitation_id": r.get("solicitation_id", ""),
            "solicitation_summary": r.get("solicitation_summary", ""),
            "procurement_officer_buyer": r.get("procurement_officer_buyer", ""),
            "contact_email": r.get("contact_email", ""),
            "additional_instructions": r.get("additional_instructions", ""),
            "procurement_program_goals": r.get("procurement_program_goals", ""),
            "status": None,         # set below
            "tags": r.get("tags",""),
            "score_bd_fit": r.get("score_bd_fit",""),
        }




        # Ensure datetime fields are Excel-safe (naive)
        row["publish_dt_et"] = to_excel_naive(row["publish_dt_et"])
        row["due_dt_et"]     = to_excel_naive(row["due_dt_et"])




        rid = row["record_id"]
        touched_ids.add(rid)




        if rid not in index:
            # NEW
            row["first_seen_et"] = ts_run_xl
            row["status"] = "New"
            ws_master.append([row[k] for k in MASTER_HDR])
            actions.append(Action("New", row))
            # update index to include this newly added row
            index[rid] = ws_master.max_row
        else:
            # EXISTING
            existing = row_dict(ws_master, index[rid])
            row["first_seen_et"] = existing.get("first_seen_et") or ts_run_xl
            if rows_equal(existing, row):
                # Unchanged
                ws_master.cell(row=index[rid], column=col_idx["last_seen_et"]).value = ts_run_xl
                ws_master.cell(row=index[rid], column=col_idx["status"]).value = "Unchanged"
                actions.append(Action("Unchanged", row))
            else:
                # Updated fields
                for k in ["url","title","agency","category","procurement_method",
                          "publish_dt_et","due_dt_et","solicitation_id","solicitation_summary",
                          "procurement_officer_buyer","contact_email","additional_instructions",
                          "procurement_program_goals","tags","score_bd_fit"]:
                    val = row[k]
                    if k in ("publish_dt_et","due_dt_et"):
                        val = to_excel_naive(val)
                    ws_master.cell(row=index[rid], column=col_idx[k]).value = val
                ws_master.cell(row=index[rid], column=col_idx["last_seen_et"]).value = ts_run_xl
                ws_master.cell(row=index[rid], column=col_idx["status"]).value = "Updated"
                actions.append(Action("Updated", row))




    # Prune stale: any record not touched this run AND last_seen_et older than STALE_AFTER_D days
    header, col_idx, index = ws_to_index(ws_master)
    stale_cutoff = to_excel_naive(ts_run - timedelta(days=STALE_AFTER_D))
    to_archive = []
    for rid, rownum in list(index.items()):
        if rid in touched_ids:
            continue
        last_seen = ws_master.cell(row=rownum, column=col_idx["last_seen_et"]).value
        if isinstance(last_seen, datetime) and last_seen < stale_cutoff:
            # mark stale and move to Archive
            ws_master.cell(row=rownum, column=col_idx["status"]).value = "Stale"
            arc_values = [ws_master.cell(row=rownum, column=i+1).value for i in range(len(MASTER_HDR))]
            ws_archive.append(arc_values)
            to_archive.append(rownum)
            actions.append(Action("Stale", row_dict(ws_master, rownum)))




    # Delete archived rows from Master (bottom-up)
    for rownum in sorted(to_archive, reverse=True):
        ws_master.delete_rows(rownum, 1)





    # Refresh style
    ensure_master_table_style(ws_master)
    auto_col_widths(ws_master)
    apply_status_conditional_formats(ws_master)





    for a in actions:
        row_for_log = a.row.copy()
        for k in ("first_seen_et","last_seen_et","publish_dt_et","due_dt_et"):
            row_for_log[k] = to_excel_naive(row_for_log.get(k))
        ws_log.append([ts_run_xl, a.action] + [row_for_log.get(k) for k in MASTER_HDR])




    # Create backup before saving
    create_workbook_backup(WORKBOOK_PATH)
    wb.save(WORKBOOK_PATH)
    logger.info(
        "Workbook saved: %s | Actions -> New:%d Updated:%d Unchanged:%d Stale:%d",
        WORKBOOK_PATH,
        sum(1 for x in actions if x.action == "New"),
        sum(1 for x in actions if x.action == "Updated"),
        sum(1 for x in actions if x.action == "Unchanged"),
        sum(1 for x in actions if x.action == "Stale"),
    )




# -------------------- Entry point --------------------
def main():
    parser = argparse.ArgumentParser(description="Scrape Maryland eMMA listings into the Excel workbook.")
    parser.add_argument("--days-ago", type=int, default=DAYS_AGO,
                        help="Days prior to today to target (default from DAYS_AGO env).")
    parser.add_argument("--skip-details", action="store_true",
                        help="Skip fetching detail pages for each listing.")
    parser.add_argument("--log-level", help="Override log level (e.g. DEBUG, INFO, WARNING).")
    args = parser.parse_args()

    if args.log_level:
        configure_logging(args.log_level)

    staging = emma_scrape(DAYS_AGO=args.days_ago, max_pages=MAX_PAGES, sleep_s=SLEEP_BETWEEN,
                          fetch_details=not args.skip_details)
    merge_into_excel(staging)
    print(f"Updated workbook: {WORKBOOK_PATH} (target day: {args.days_ago} days ago)")


if __name__ == "__main__":
    main()
