"""
Ultimate eMMA Scraper with Enhanced Data Accuracy and Deduplication

Features:
- Zero data duplication with multi-level deduplication
- Maximum information extraction
- Improved, meaningful column names
- Data validation and cleaning
- Comprehensive field extraction
"""

import argparse
import os
import re
import time
import logging
import json
import csv
import hashlib
from hashlib import blake2b, sha256
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from urllib.parse import urljoin, urlparse, parse_qs
from zipfile import BadZipFile
from typing import Optional, Tuple, List, Dict, Any, Set
import shutil
from pathlib import Path
from collections import defaultdict

import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup, NavigableString
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

# -------------------- Configuration --------------------
def get_default_workbook_path():
    """Get platform-appropriate default workbook path."""
    if os.name == 'nt':  # Windows
        return r"C:\Users\hkhoshhal001\Guidehouse\New York Mid-Atlantic (NYMA) - 05. Scanning resources\Automated Scanning\opportunities.xlsx"
    else:  # Unix/Linux/Mac
        home = os.path.expanduser("~")
        return os.path.join(home, "Documents", "emma_opportunities.xlsx")

WORKBOOK_PATH = os.getenv("EMMA_XLSX", get_default_workbook_path())
DAYS_AGO = int(os.getenv("DAYS_AGO", "0"))
STALE_AFTER_D = int(os.getenv("STALE_AFTER_D", "7"))
MAX_PAGES = int(os.getenv("MAX_PAGES", "50"))
SLEEP_BETWEEN = float(os.getenv("SLEEP_BETWEEN", "1.0"))
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
TIMEOUT_SECONDS = int(os.getenv("TIMEOUT_SECONDS", "30"))
USER_AGENT = os.getenv("USER_AGENT", "Mozilla/5.0 (compatible; MD-EmmaScraper/2.0)")

BASE = "https://emma.maryland.gov"
BROWSE_URL = f"{BASE}/page.aspx/en/rfp/request_browse_public"

# -------------------- Timezone --------------------
try:
    from zoneinfo import ZoneInfo
    ET_TZ = ZoneInfo("America/New_York")
except Exception:
    ET_TZ = None

def now_et():
    return datetime.now(ET_TZ) if ET_TZ else datetime.now()

def localize_et(dt: datetime) -> datetime:
    return dt.replace(tzinfo=ET_TZ) if ET_TZ and dt else dt

def to_excel_naive(dt):
    """Excel cannot store tz-aware datetimes."""
    if dt is None:
        return None
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt

# -------------------- Logging --------------------
logger = logging.getLogger("emma_scraper_ultimate")

def configure_logging(level: str = "INFO"):
    """Configure logging."""
    logger.setLevel(getattr(logging, level))
    handler = logging.StreamHandler()
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)

# -------------------- Enhanced Column Schema --------------------
# More meaningful and comprehensive column names
ENHANCED_COLUMNS = [
    # Identification
    "data_source",           # Source system (emma)
    "unique_id",            # Our unique record ID
    "solicitation_number",   # Official solicitation/RFP number
    "emma_id",              # eMMA system ID (extracted from URL)

    # Basic Information
    "opportunity_title",     # Full title/description
    "issuing_agency",       # Agency/department name
    "category",             # Procurement category
    "procurement_type",     # Method/type of procurement

    # Dates and Timing
    "published_date",       # When posted
    "response_deadline",    # When due
    "days_until_due",       # Calculated days remaining
    "pre_bid_conference",   # Pre-bid meeting date if any

    # Contact Information
    "buyer_name",           # Procurement officer/buyer
    "contact_email",        # Primary email
    "contact_phone",        # Primary phone
    "contact_fax",          # Fax if available
    "contact_address",      # Physical address if available

    # Detailed Information
    "project_description",  # Detailed description/summary
    "submission_instructions", # How to submit
    "special_requirements", # Special instructions/requirements
    "small_business_goals", # MBE/WBE/SBE goals
    "estimated_value",      # Contract value if disclosed
    "contract_duration",    # Duration/period of performance
    "incumbent_vendor",     # Current vendor if renewal

    # Documents and Links
    "opportunity_url",      # Direct link to opportunity
    "attachments_count",    # Number of attachments
    "attachment_names",     # List of attachment filenames
    "amendment_count",      # Number of amendments
    "q_and_a_deadline",     # Question submission deadline

    # Metadata and Tracking
    "first_seen_date",      # When we first saw it
    "last_updated_date",    # When we last checked
    "data_status",          # New/Updated/Unchanged/Stale
    "change_history",       # JSON of what changed
    "validation_flags",     # Data quality flags

    # Analysis Fields
    "auto_tags",            # System-generated tags
    "relevance_score",      # Calculated relevance
    "priority_level",       # High/Medium/Low
    "notes",               # Manual notes field
]

# Map old column names to new meaningful names
COLUMN_MAPPING = {
    "source": "data_source",
    "record_id": "unique_id",
    "solicitation_id": "solicitation_number",
    "url": "opportunity_url",
    "title": "opportunity_title",
    "agency": "issuing_agency",
    "procurement_method": "procurement_type",
    "publish_dt_et": "published_date",
    "due_dt_et": "response_deadline",
    "solicitation_summary": "project_description",
    "procurement_officer_buyer": "buyer_name",
    "additional_instructions": "submission_instructions",
    "procurement_program_goals": "small_business_goals",
    "first_seen_et": "first_seen_date",
    "last_seen_et": "last_updated_date",
    "status": "data_status",
    "tags": "auto_tags",
    "score_bd_fit": "relevance_score",
}

# -------------------- Data Validation and Cleaning --------------------
class DataValidator:
    """Validate and clean extracted data."""

    @staticmethod
    def clean_text(text: str) -> str:
        """Clean and normalize text."""
        if not text:
            return ""

        # Remove excessive whitespace
        text = re.sub(r'\s+', ' ', text)
        # Remove non-printable characters
        text = ''.join(char for char in text if char.isprintable() or char in '\n\r\t')
        # Trim
        text = text.strip()

        return text

    @staticmethod
    def validate_email(email: str) -> str:
        """Validate and clean email."""
        if not email:
            return ""

        email = email.lower().strip()
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return email
        return ""

    @staticmethod
    def validate_phone(phone: str) -> str:
        """Validate and format phone number."""
        if not phone:
            return ""

        # Extract digits
        digits = re.sub(r'\D', '', phone)

        # Format if valid US phone
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            return f"+1 ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"

        return phone.strip()

    @staticmethod
    def validate_date(date_val) -> Optional[datetime]:
        """Validate date value."""
        if isinstance(date_val, datetime):
            return date_val
        if isinstance(date_val, str):
            # Try parsing
            for fmt in [
                "%m/%d/%Y %I:%M:%S %p",
                "%m/%d/%Y %I:%M %p",
                "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d",
            ]:
                try:
                    return datetime.strptime(date_val, fmt)
                except ValueError:
                    continue
        return None

    @staticmethod
    def extract_money_value(text: str) -> str:
        """Extract monetary value from text."""
        if not text:
            return ""

        # Look for dollar amounts
        patterns = [
            r'\$[\d,]+(?:\.\d{2})?(?:\s*(?:million|billion|M|B))?',
            r'[\d,]+(?:\.\d{2})?\s*(?:dollars|USD)',
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(0)

        return ""

# -------------------- Enhanced Deduplication System --------------------
class DeduplicationManager:
    """Multi-level deduplication to ensure zero duplicates."""

    def __init__(self):
        self.seen_hashes = set()
        self.seen_ids = set()
        self.seen_urls = set()
        self.seen_composite_keys = set()

    def create_composite_key(self, row: dict) -> str:
        """Create composite key for deduplication."""
        # Primary key: solicitation number if available
        if row.get("solicitation_number"):
            return f"sol_{row['solicitation_number'].lower().strip()}"

        # Secondary: URL-based ID
        if row.get("emma_id"):
            return f"emma_{row['emma_id']}"

        # Tertiary: Title + Agency + Date
        title = (row.get("opportunity_title") or "").lower().strip()
        agency = (row.get("issuing_agency") or "").lower().strip()
        date = row.get("published_date")

        if title and agency and date:
            date_str = date.isoformat() if isinstance(date, datetime) else str(date)
            composite = f"{title}|{agency}|{date_str}"
            return f"comp_{hashlib.sha256(composite.encode()).hexdigest()[:16]}"

        # Fallback: Full row hash
        row_str = json.dumps(row, sort_keys=True, default=str)
        return f"hash_{hashlib.sha256(row_str.encode()).hexdigest()[:16]}"

    def is_duplicate(self, row: dict) -> bool:
        """Check if row is duplicate using multiple strategies."""
        # Check solicitation number
        sol_num = row.get("solicitation_number")
        if sol_num and sol_num in self.seen_ids:
            logger.debug(f"Duplicate found by solicitation number: {sol_num}")
            return True

        # Check URL
        url = row.get("opportunity_url")
        if url and url in self.seen_urls:
            logger.debug(f"Duplicate found by URL: {url}")
            return True

        # Check composite key
        comp_key = self.create_composite_key(row)
        if comp_key in self.seen_composite_keys:
            logger.debug(f"Duplicate found by composite key: {comp_key}")
            return True

        # Check content hash
        content = f"{row.get('opportunity_title')}|{row.get('project_description')}"
        content_hash = hashlib.sha256(content.encode()).hexdigest()
        if content_hash in self.seen_hashes:
            logger.debug(f"Duplicate found by content hash")
            return True

        return False

    def mark_seen(self, row: dict):
        """Mark row as seen."""
        if row.get("solicitation_number"):
            self.seen_ids.add(row["solicitation_number"])

        if row.get("opportunity_url"):
            self.seen_urls.add(row["opportunity_url"])

        comp_key = self.create_composite_key(row)
        self.seen_composite_keys.add(comp_key)

        content = f"{row.get('opportunity_title')}|{row.get('project_description')}"
        self.seen_hashes.add(hashlib.sha256(content.encode()).hexdigest())

# -------------------- Enhanced Field Extraction --------------------
class FieldExtractor:
    """Extract maximum information from pages."""

    def __init__(self):
        self.validator = DataValidator()

    def extract_all_fields(self, soup: BeautifulSoup, base_url: str = BASE) -> dict:
        """Extract all possible fields from a page."""
        data = {}

        # Extract from tables
        for table in soup.find_all("table"):
            self._extract_from_table(table, data)

        # Extract from definition lists
        for dl in soup.find_all("dl"):
            self._extract_from_dl(dl, data)

        # Extract from labeled divs
        self._extract_from_labeled_elements(soup, data)

        # Extract specific patterns
        self._extract_patterns(soup, data)

        # Extract attachments
        self._extract_attachments(soup, data, base_url)

        # Clean all text fields
        for key in data:
            if isinstance(data[key], str):
                data[key] = self.validator.clean_text(data[key])

        return data

    def _extract_from_table(self, table, data: dict):
        """Extract key-value pairs from table."""
        for row in table.find_all("tr"):
            cells = row.find_all(["th", "td"])
            if len(cells) >= 2:
                label = self.validator.clean_text(cells[0].get_text()).lower()
                value = self.validator.clean_text(cells[1].get_text())

                if not label or not value:
                    continue

                # Map to our fields
                self._map_field(label, value, data)

    def _extract_from_dl(self, dl, data: dict):
        """Extract from definition list."""
        dt_elements = dl.find_all("dt")
        dd_elements = dl.find_all("dd")

        for dt, dd in zip(dt_elements, dd_elements):
            label = self.validator.clean_text(dt.get_text()).lower()
            value = self.validator.clean_text(dd.get_text())

            if label and value:
                self._map_field(label, value, data)

    def _extract_from_labeled_elements(self, soup, data: dict):
        """Extract from elements with labels."""
        # Look for label-value patterns
        for element in soup.find_all(["div", "span", "p"]):
            text = element.get_text()

            # Pattern: "Label: Value"
            match = re.search(r'^([^:]+):\s*(.+)$', text)
            if match:
                label = self.validator.clean_text(match.group(1)).lower()
                value = self.validator.clean_text(match.group(2))

                if label and value:
                    self._map_field(label, value, data)

    def _extract_patterns(self, soup, data: dict):
        """Extract specific patterns from text."""
        full_text = soup.get_text()

        # Email addresses
        emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', full_text)
        if emails and not data.get("contact_email"):
            data["contact_email"] = self.validator.validate_email(emails[0])

        # Phone numbers
        phones = re.findall(r'(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', full_text)
        if phones and not data.get("contact_phone"):
            data["contact_phone"] = self.validator.validate_phone(phones[0])

        # Dollar amounts
        if not data.get("estimated_value"):
            value = self.validator.extract_money_value(full_text)
            if value:
                data["estimated_value"] = value

        # Dates
        date_patterns = [
            r'\d{1,2}/\d{1,2}/\d{4}',
            r'\d{4}-\d{2}-\d{2}',
        ]
        for pattern in date_patterns:
            dates = re.findall(pattern, full_text)
            if dates and not data.get("response_deadline"):
                # Try to find due date
                for date_str in dates:
                    if any(keyword in full_text[max(0, full_text.index(date_str)-50):full_text.index(date_str)]
                           for keyword in ["due", "deadline", "close", "submit"]):
                        data["response_deadline"] = date_str
                        break

    def _extract_attachments(self, soup, data: dict, base_url: str):
        """Extract attachment information."""
        attachments = []

        # Look for download links
        for link in soup.find_all("a", href=True):
            href = link["href"]
            text = link.get_text()

            if any(ext in href.lower() for ext in ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.zip']):
                attachments.append({
                    "name": self.validator.clean_text(text),
                    "url": urljoin(base_url, href)
                })

        if attachments:
            data["attachments_count"] = len(attachments)
            data["attachment_names"] = "; ".join([a["name"] for a in attachments])

    def _map_field(self, label: str, value: str, data: dict):
        """Map extracted field to our schema."""
        # Normalize label
        label = label.lower().strip()

        # Direct mappings
        mappings = {
            "solicitation number": "solicitation_number",
            "solicitation #": "solicitation_number",
            "rfp #": "solicitation_number",
            "bid #": "solicitation_number",
            "title": "opportunity_title",
            "description": "project_description",
            "summary": "project_description",
            "agency": "issuing_agency",
            "department": "issuing_agency",
            "buyer": "buyer_name",
            "procurement officer": "buyer_name",
            "contact": "buyer_name",
            "email": "contact_email",
            "phone": "contact_phone",
            "telephone": "contact_phone",
            "fax": "contact_fax",
            "address": "contact_address",
            "due date": "response_deadline",
            "closing date": "response_deadline",
            "deadline": "response_deadline",
            "published": "published_date",
            "posted": "published_date",
            "issue date": "published_date",
            "pre-bid": "pre_bid_conference",
            "pre bid": "pre_bid_conference",
            "conference": "pre_bid_conference",
            "value": "estimated_value",
            "amount": "estimated_value",
            "duration": "contract_duration",
            "period": "contract_duration",
            "incumbent": "incumbent_vendor",
            "current vendor": "incumbent_vendor",
            "mbe": "small_business_goals",
            "wbe": "small_business_goals",
            "sbe": "small_business_goals",
            "small business": "small_business_goals",
            "instruction": "submission_instructions",
            "how to": "submission_instructions",
            "requirement": "special_requirements",
            "q&a": "q_and_a_deadline",
            "question": "q_and_a_deadline",
        }

        # Find best match
        for key_pattern, field_name in mappings.items():
            if key_pattern in label:
                # Don't overwrite with empty values
                if value and (not data.get(field_name) or len(value) > len(data.get(field_name, ""))):
                    data[field_name] = value
                break

# -------------------- Enhanced Scraping Functions --------------------
def make_session() -> requests.Session:
    """Create HTTP session."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    })

    retries = Retry(
        total=5,
        backoff_factor=1.0,
        status_forcelist=[403, 429, 500, 502, 503, 504]
    )

    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.mount("http://", HTTPAdapter(max_retries=retries))
    return s

def parse_hidden_fields(soup: BeautifulSoup) -> dict:
    """Parse ASP.NET hidden fields."""
    fields = {}
    for name in ["__VIEWSTATE", "__EVENTVALIDATION", "__VIEWSTATEGENERATOR", "__EVENTTARGET", "__EVENTARGUMENT"]:
        el = soup.find("input", {"name": name})
        if el and el.has_attr("value"):
            fields[name] = el["value"]
    return fields

def extract_emma_id(url: str) -> str:
    """Extract eMMA ID from URL."""
    if not url:
        return ""

    # Pattern: /extranet/123456
    match = re.search(r'/extranet/(\d+)', url)
    if match:
        return match.group(1)

    # Pattern: requestId=123456
    parsed = urlparse(url)
    params = parse_qs(parsed.query)
    if 'requestId' in params:
        return params['requestId'][0]

    return ""

def scrape_emma_enhanced() -> List[dict]:
    """Enhanced scraping with maximum extraction and zero duplication."""
    session = make_session()
    dedup_manager = DeduplicationManager()
    field_extractor = FieldExtractor()
    validator = DataValidator()

    all_records = []

    try:
        # Get initial page
        response = session.get(BROWSE_URL, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        pages_scraped = 0

        while pages_scraped < MAX_PAGES:
            # Find results table
            table = soup.find("table", class_="iv-grid-view") or soup.find("table")

            if not table:
                logger.warning("No results table found")
                break

            # Extract rows
            rows = table.find_all("tr")[1:]  # Skip header

            for row in rows:
                cells = row.find_all("td")
                if not cells:
                    continue

                record = {}

                # Extract from listing page
                # Assuming standard column order (adjust based on actual structure)
                if len(cells) > 0:
                    # Title and URL
                    title_cell = cells[0]
                    link = title_cell.find("a")
                    if link:
                        record["opportunity_title"] = validator.clean_text(link.get_text())
                        record["opportunity_url"] = urljoin(BASE, link.get("href", ""))
                        record["emma_id"] = extract_emma_id(record["opportunity_url"])
                    else:
                        record["opportunity_title"] = validator.clean_text(title_cell.get_text())

                # Extract other columns (adjust indices based on actual structure)
                if len(cells) > 1:
                    record["category"] = validator.clean_text(cells[1].get_text())
                if len(cells) > 2:
                    record["procurement_type"] = validator.clean_text(cells[2].get_text())
                if len(cells) > 3:
                    record["issuing_agency"] = validator.clean_text(cells[3].get_text())
                if len(cells) > 4:
                    date_text = cells[4].get_text()
                    record["published_date"] = validator.validate_date(date_text)

                # Skip if duplicate
                if dedup_manager.is_duplicate(record):
                    logger.debug(f"Skipping duplicate: {record.get('opportunity_title')}")
                    continue

                # Fetch detail page for more information
                if record.get("opportunity_url"):
                    try:
                        detail_response = session.get(record["opportunity_url"], timeout=TIMEOUT_SECONDS)
                        detail_soup = BeautifulSoup(detail_response.text, 'html.parser')

                        # Extract all available fields
                        detail_data = field_extractor.extract_all_fields(detail_soup, BASE)

                        # Merge detail data (don't overwrite existing non-empty values)
                        for key, value in detail_data.items():
                            if value and not record.get(key):
                                record[key] = value

                        time.sleep(SLEEP_BETWEEN)

                    except Exception as e:
                        logger.warning(f"Failed to fetch details for {record.get('opportunity_url')}: {e}")

                # Generate unique ID
                record["unique_id"] = dedup_manager.create_composite_key(record)

                # Set metadata
                record["data_source"] = "emma"
                record["first_seen_date"] = to_excel_naive(now_et())
                record["last_updated_date"] = to_excel_naive(now_et())
                record["data_status"] = "New"

                # Calculate days until due
                if record.get("response_deadline") and record.get("published_date"):
                    try:
                        due = validator.validate_date(record["response_deadline"])
                        pub = validator.validate_date(record["published_date"])
                        if due and pub:
                            delta = (due - pub).days
                            record["days_until_due"] = delta
                    except:
                        pass

                # Mark as seen and add to results
                dedup_manager.mark_seen(record)
                all_records.append(record)

                logger.info(f"Extracted: {record.get('opportunity_title', 'Unknown')[:50]}...")

            # Find next page
            next_link = None
            for link in soup.find_all("a", href=True):
                if "next" in link.get_text().lower() or "›" in link.get_text():
                    href = link["href"]
                    if "__doPostBack" in href:
                        match = re.search(r"__doPostBack\('([^']+)','([^']+)'\)", href)
                        if match:
                            next_link = match.groups()
                            break

            if not next_link:
                logger.info("No more pages")
                break

            # Go to next page
            fields = parse_hidden_fields(soup)
            fields["__EVENTTARGET"] = next_link[0]
            fields["__EVENTARGUMENT"] = next_link[1]

            response = session.post(BROWSE_URL, data=fields, timeout=TIMEOUT_SECONDS)
            soup = BeautifulSoup(response.text, 'html.parser')

            pages_scraped += 1
            logger.info(f"Scraped page {pages_scraped + 1}")
            time.sleep(SLEEP_BETWEEN)

    except Exception as e:
        logger.error(f"Scraping failed: {e}")

    logger.info(f"Total unique records extracted: {len(all_records)}")

    # Final validation pass
    validated_records = []
    for record in all_records:
        # Ensure all expected columns exist
        for col in ENHANCED_COLUMNS:
            if col not in record:
                record[col] = ""

        # Validate specific fields
        if record.get("contact_email"):
            record["contact_email"] = validator.validate_email(record["contact_email"])
        if record.get("contact_phone"):
            record["contact_phone"] = validator.validate_phone(record["contact_phone"])

        validated_records.append(record)

    return validated_records

# -------------------- Excel Operations --------------------
def save_to_excel(records: List[dict], filepath: str):
    """Save records to Excel with enhanced columns."""
    try:
        # Try to load existing workbook
        try:
            wb = load_workbook(filepath)
        except:
            wb = Workbook()

        # Get or create Master sheet
        if "Master" in wb.sheetnames:
            ws = wb["Master"]
            # Clear existing data
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.active
            ws.title = "Master"

        # Write headers
        ws.append(ENHANCED_COLUMNS)

        # Write data
        for record in records:
            row = []
            for col in ENHANCED_COLUMNS:
                value = record.get(col, "")
                # Convert datetime for Excel
                if isinstance(value, datetime):
                    value = to_excel_naive(value)
                row.append(value)
            ws.append(row)

        # Apply formatting
        apply_excel_formatting(ws)

        # Create backup
        if os.path.exists(filepath):
            backup_path = filepath.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            shutil.copy2(filepath, backup_path)
            logger.info(f"Created backup: {backup_path}")

        # Save
        wb.save(filepath)
        logger.info(f"Saved {len(records)} records to {filepath}")

        # Create summary report
        create_summary_report(records, wb)
        wb.save(filepath)

    except Exception as e:
        logger.error(f"Failed to save to Excel: {e}")
        raise

def apply_excel_formatting(ws):
    """Apply formatting to Excel sheet."""
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add table
    if ws.max_row > 1:
        table = Table(displayName="OpportunitiesTable",
                     ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9",
                              showFirstColumn=False,
                              showLastColumn=False,
                              showRowStripes=True,
                              showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

def create_summary_report(records: List[dict], wb):
    """Create summary analytics sheet."""
    # Create or get Summary sheet
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Summary")

    # Calculate statistics
    total = len(records)
    agencies = defaultdict(int)
    categories = defaultdict(int)

    for record in records:
        agencies[record.get("issuing_agency", "Unknown")] += 1
        categories[record.get("category", "Unknown")] += 1

    # Write summary
    ws.append(["eMMA Opportunities Summary Report"])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    ws.append(["Total Opportunities:", total])
    ws.append([])

    ws.append(["By Agency:"])
    for agency, count in sorted(agencies.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws.append(["", agency, count])
    ws.append([])

    ws.append(["By Category:"])
    for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws.append(["", category, count])

# -------------------- Main --------------------
def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Ultimate eMMA Scraper")
    parser.add_argument("--output", default=WORKBOOK_PATH, help="Output Excel file")
    parser.add_argument("--log-level", default=LOG_LEVEL, help="Log level")

    args = parser.parse_args()

    configure_logging(args.log_level)

    logger.info("Starting enhanced eMMA scraping...")

    # Scrape data
    records = scrape_emma_enhanced()

    if records:
        # Save to Excel
        save_to_excel(records, args.output)

        # Print summary
        logger.info(f"\n{'='*50}")
        logger.info(f"Scraping Complete!")
        logger.info(f"Total unique records: {len(records)}")
        logger.info(f"Output saved to: {args.output}")
        logger.info(f"{'='*50}")
    else:
        logger.warning("No records extracted")

if __name__ == "__main__":
    main()