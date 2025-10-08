"""
Fixed eMMA Scraper with Correct Table Structure
"""

import argparse
import os
import re
import time
import logging
import json
import hashlib
from datetime import datetime, timedelta
from urllib.parse import urljoin
from typing import Optional, List, Dict
import shutil

import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Configuration
WORKBOOK_PATH = os.getenv("EMMA_XLSX", "opportunities.xlsx")
MAX_PAGES = int(os.getenv("MAX_PAGES", "1"))
SLEEP_BETWEEN = float(os.getenv("SLEEP_BETWEEN", "0.5"))
TIMEOUT = 30

BASE = "https://emma.maryland.gov"
BROWSE_URL = f"{BASE}/page.aspx/en/rfp/request_browse_public"

# Logger
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Enhanced column schema with meaningful names
COLUMNS = [
    "unique_id",
    "solicitation_id",
    "opportunity_title",
    "status",
    "response_deadline",
    "published_date",
    "main_category",
    "solicitation_type",
    "issuing_agency",
    "procurement_officer",
    "sub_agency",
    "emma_system_id",
    "opportunity_url",
    "days_until_due",
    "first_seen",
    "last_updated",
    "data_quality_score",
]

class DataCleaner:
    """Clean and validate data."""

    @staticmethod
    def clean_text(text):
        """Clean text data."""
        if not text:
            return ""
        # Remove excessive whitespace
        text = re.sub(r'\s+', ' ', str(text).strip())
        return text

    @staticmethod
    def parse_date(date_str):
        """Parse date string."""
        if not date_str:
            return None

        # Clean the string
        date_str = DataCleaner.clean_text(date_str)

        # Try different formats
        formats = [
            "%m/%d/%Y",
            "%m/%d/%Y %I:%M:%S %p",
            "%m/%d/%Y %H:%M:%S",
            "%Y-%m-%d",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str.split()[0] if ' ' in date_str else date_str, fmt)
            except:
                continue

        return None

class DuplicateChecker:
    """Check for duplicate records."""

    def __init__(self):
        self.seen_ids = set()
        self.seen_titles = set()
        self.seen_hashes = set()

    def is_duplicate(self, record):
        """Check if record is duplicate."""
        # Check by solicitation ID
        sol_id = record.get("solicitation_id")
        if sol_id and sol_id in self.seen_ids:
            return True

        # Check by title + agency combo
        title = record.get("opportunity_title", "")
        agency = record.get("issuing_agency", "")
        combo = f"{title}|{agency}".lower()
        combo_hash = hashlib.md5(combo.encode()).hexdigest()

        if combo_hash in self.seen_hashes:
            return True

        return False

    def mark_seen(self, record):
        """Mark record as seen."""
        sol_id = record.get("solicitation_id")
        if sol_id:
            self.seen_ids.add(sol_id)

        title = record.get("opportunity_title", "")
        agency = record.get("issuing_agency", "")
        combo = f"{title}|{agency}".lower()
        combo_hash = hashlib.md5(combo.encode()).hexdigest()
        self.seen_hashes.add(combo_hash)

def make_session():
    """Create HTTP session."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })

    retry = Retry(total=3, backoff_factor=0.5)
    session.mount("https://", HTTPAdapter(max_retries=retry))

    return session

def extract_records_from_page(soup):
    """Extract records from the page."""
    records = []
    cleaner = DataCleaner()

    # Find the main table
    table = soup.find("table", class_="iv-grid-view")
    if not table:
        logger.warning("No results table found")
        return records

    # Get all rows (skip header)
    rows = table.find_all("tr")[1:]

    for row in rows:
        cells = row.find_all("td")
        if len(cells) < 8:
            continue

        record = {}

        # Map based on actual column positions from eMMA
        # 0: Edit link
        # 1: ID (BPM number)
        # 2: Title
        # 3: Status
        # 4: Due/Close Date
        # 5: Publish Date
        # 6: Main Category
        # 7: Solicitation Type
        # 8: Issuing Agency
        # 12: Procurement Officer (if exists)
        # 17: eMM ID (if exists)

        # Extract ID
        if len(cells) > 1:
            record["solicitation_id"] = cleaner.clean_text(cells[1].get_text())

        # Extract Title and URL
        if len(cells) > 2:
            title_cell = cells[2]
            record["opportunity_title"] = cleaner.clean_text(title_cell.get_text())

            # Try to get URL from edit link in first cell
            if len(cells) > 0:
                edit_cell = cells[0]
                link = edit_cell.find("a")
                if link and link.get("href"):
                    record["opportunity_url"] = urljoin(BASE, link["href"])

        # Extract Status
        if len(cells) > 3:
            record["status"] = cleaner.clean_text(cells[3].get_text())

        # Extract Due Date
        if len(cells) > 4:
            due_text = cleaner.clean_text(cells[4].get_text())
            record["response_deadline"] = due_text
            parsed_due = cleaner.parse_date(due_text)
            if parsed_due:
                days_until = (parsed_due - datetime.now()).days
                record["days_until_due"] = days_until

        # Extract Publish Date
        if len(cells) > 5:
            pub_text = cleaner.clean_text(cells[5].get_text())
            record["published_date"] = pub_text

        # Extract Category
        if len(cells) > 6:
            record["main_category"] = cleaner.clean_text(cells[6].get_text())

        # Extract Solicitation Type
        if len(cells) > 7:
            record["solicitation_type"] = cleaner.clean_text(cells[7].get_text())

        # Extract Issuing Agency
        if len(cells) > 8:
            record["issuing_agency"] = cleaner.clean_text(cells[8].get_text())

        # Extract Procurement Officer if available
        if len(cells) > 12:
            record["procurement_officer"] = cleaner.clean_text(cells[12].get_text())

        # Extract Sub Agency if available
        if len(cells) > 14:
            record["sub_agency"] = cleaner.clean_text(cells[14].get_text())

        # Extract eMM ID if available
        if len(cells) > 17:
            record["emma_system_id"] = cleaner.clean_text(cells[17].get_text())

        # Generate unique ID
        if record.get("solicitation_id"):
            record["unique_id"] = f"sol_{record['solicitation_id']}"
        else:
            # Fallback to hash of title+agency
            combo = f"{record.get('opportunity_title', '')}|{record.get('issuing_agency', '')}"
            record["unique_id"] = f"hash_{hashlib.md5(combo.encode()).hexdigest()[:12]}"

        # Add metadata
        record["first_seen"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        record["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Calculate data quality score
        filled_fields = sum(1 for v in record.values() if v and str(v).strip())
        total_fields = len(record)
        record["data_quality_score"] = round((filled_fields / total_fields) * 100, 1)

        records.append(record)

    return records

def scrape_emma():
    """Main scraping function."""
    session = make_session()
    duplicate_checker = DuplicateChecker()
    all_records = []

    try:
        # Get initial page
        response = session.get(BROWSE_URL, timeout=TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract records from first page
        records = extract_records_from_page(soup)

        for record in records:
            if not duplicate_checker.is_duplicate(record):
                duplicate_checker.mark_seen(record)
                all_records.append(record)
                logger.info(f"Extracted: {record.get('opportunity_title', 'Unknown')[:60]}...")
            else:
                logger.debug(f"Skipped duplicate: {record.get('opportunity_title', '')}")

        # For now, just scrape the first page (can be extended for pagination)

    except Exception as e:
        logger.error(f"Scraping failed: {e}")

    return all_records

def save_to_excel(records, filepath):
    """Save records to Excel."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Opportunities"

        # Write headers
        ws.append(COLUMNS)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for record in records:
            row = []
            for col in COLUMNS:
                value = record.get(col, "")
                row.append(value)
            ws.append(row)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add table formatting
        if ws.max_row > 1:
            tab = Table(displayName="OpportunitiesTable",
                       ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            style = TableStyleInfo(name="TableStyleMedium9",
                                  showFirstColumn=False,
                                  showLastColumn=False,
                                  showRowStripes=True,
                                  showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

        # Freeze header
        ws.freeze_panes = "A2"

        # Add summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["eMMA Opportunities Summary"])
        summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        summary.append([])
        summary.append(["Total Opportunities:", len(records)])
        summary.append([])

        # Count by agency
        agencies = {}
        categories = {}
        for record in records:
            agency = record.get("issuing_agency", "Unknown")
            category = record.get("main_category", "Unknown")
            agencies[agency] = agencies.get(agency, 0) + 1
            categories[category] = categories.get(category, 0) + 1

        summary.append(["Top Agencies:"])
        for agency, count in sorted(agencies.items(), key=lambda x: x[1], reverse=True)[:5]:
            summary.append(["", agency, count])

        summary.append([])
        summary.append(["Top Categories:"])
        for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:5]:
            summary.append(["", category, count])

        # Save
        wb.save(filepath)
        logger.info(f"Saved {len(records)} records to {filepath}")

        return True

    except Exception as e:
        logger.error(f"Failed to save Excel: {e}")
        return False

def print_sample_data(records):
    """Print sample of extracted data."""
    print("\n" + "="*80)
    print("SAMPLE OF EXTRACTED DATA")
    print("="*80)

    for i, record in enumerate(records[:3], 1):
        print(f"\nRecord {i}:")
        print("-" * 40)
        print(f"  Unique ID: {record.get('unique_id')}")
        print(f"  Solicitation ID: {record.get('solicitation_id')}")
        print(f"  Title: {record.get('opportunity_title', '')[:80]}")
        print(f"  Agency: {record.get('issuing_agency')}")
        print(f"  Category: {record.get('main_category')}")
        print(f"  Type: {record.get('solicitation_type')}")
        print(f"  Status: {record.get('status')}")
        print(f"  Due Date: {record.get('response_deadline')}")
        print(f"  Days Until Due: {record.get('days_until_due')}")
        print(f"  Published: {record.get('published_date')}")
        print(f"  Officer: {record.get('procurement_officer')}")
        print(f"  Data Quality: {record.get('data_quality_score')}%")

    print("\n" + "="*80)
    print(f"TOTAL UNIQUE RECORDS: {len(records)}")
    print("="*80)

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Fixed eMMA Scraper")
    parser.add_argument("--output", default=WORKBOOK_PATH, help="Output Excel file")

    args = parser.parse_args()

    logger.info("Starting eMMA scraping...")

    # Scrape data
    records = scrape_emma()

    if records:
        # Print sample
        print_sample_data(records)

        # Save to Excel
        save_to_excel(records, args.output)

        print(f"\n✅ Successfully extracted {len(records)} unique opportunities")
        print(f"📁 Output saved to: {args.output}")
    else:
        logger.warning("No records extracted")

if __name__ == "__main__":
    main()