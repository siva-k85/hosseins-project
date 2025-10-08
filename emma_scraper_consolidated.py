"""
eMMA Scraper with Consolidated Additional Information Column
Combines detailed extraction with cleaner column structure
"""

import argparse
import os
import re
import time
import logging
import json
import hashlib
from datetime import datetime, timedelta
from urllib.parse import urljoin, urlparse, parse_qs
from typing import Optional, List, Dict
import shutil

import requests
from requests.adapters import HTTPAdapter, Retry
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Configuration
WORKBOOK_PATH = os.getenv("EMMA_XLSX", "opportunities.xlsx")
MAX_PAGES = int(os.getenv("MAX_PAGES", "1"))
SLEEP_BETWEEN = float(os.getenv("SLEEP_BETWEEN", "1.0"))
TIMEOUT = 30

BASE = "https://emma.maryland.gov"
BROWSE_URL = f"{BASE}/page.aspx/en/rfp/request_browse_public"

# Logger
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Main columns - keeping the essential ones separate
MAIN_COLUMNS = [
    "unique_id",
    "solicitation_number",
    "opportunity_title",
    "issuing_agency",
    "category",
    "procurement_type",
    "status",
    "published_date",
    "response_deadline",
    "days_until_due",
    "procurement_officer",
    "contact_email",
    "contact_phone",
    "opportunity_url",
    "additional_information",  # Consolidated field for all extra details
    "first_seen",
    "last_updated",
    "data_quality_score",
]

class DataCleaner:
    """Clean and validate data."""

    @staticmethod
    def clean_text(text):
        """Clean text data."""
        if not text:
            return ""
        # Remove excessive whitespace
        text = re.sub(r'\s+', ' ', str(text).strip())
        # Remove non-printable characters except newlines
        text = ''.join(char for char in text if char.isprintable() or char in '\n\r\t')
        return text.strip()

    @staticmethod
    def parse_date(date_str):
        """Parse date string."""
        if not date_str:
            return None

        date_str = DataCleaner.clean_text(date_str)

        formats = [
            "%m/%d/%Y",
            "%m/%d/%Y %I:%M:%S %p",
            "%m/%d/%Y %H:%M:%S",
            "%Y-%m-%d",
            "%m-%d-%Y",
            "%d/%m/%Y",
        ]

        for fmt in formats:
            try:
                # Try to parse just the date part if there's time included
                date_part = date_str.split()[0] if ' ' in date_str else date_str
                return datetime.strptime(date_part, fmt)
            except:
                try:
                    return datetime.strptime(date_str, fmt)
                except:
                    continue

        return None

    @staticmethod
    def validate_email(email):
        """Validate and clean email."""
        if not email:
            return ""
        email = email.lower().strip()
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return email
        return ""

    @staticmethod
    def validate_phone(phone):
        """Format phone number."""
        if not phone:
            return ""

        # Extract digits
        digits = re.sub(r'\D', '', phone)

        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            return f"+1 ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"

        return phone.strip()

class DuplicateChecker:
    """Advanced duplicate detection."""

    def __init__(self):
        self.seen_ids = set()
        self.seen_urls = set()
        self.seen_hashes = set()

    def is_duplicate(self, record):
        """Check if record is duplicate using multiple methods."""
        # Check by solicitation ID
        sol_id = record.get("solicitation_number")
        if sol_id and sol_id in self.seen_ids:
            return True

        # Check by URL
        url = record.get("opportunity_url")
        if url and url in self.seen_urls:
            return True

        # Check by content hash (title + agency + date)
        title = record.get("opportunity_title", "")
        agency = record.get("issuing_agency", "")
        date = str(record.get("published_date", ""))
        combo = f"{title}|{agency}|{date}".lower()
        combo_hash = hashlib.md5(combo.encode()).hexdigest()

        if combo_hash in self.seen_hashes:
            return True

        return False

    def mark_seen(self, record):
        """Mark record as seen."""
        sol_id = record.get("solicitation_number")
        if sol_id:
            self.seen_ids.add(sol_id)

        url = record.get("opportunity_url")
        if url:
            self.seen_urls.add(url)

        title = record.get("opportunity_title", "")
        agency = record.get("issuing_agency", "")
        date = str(record.get("published_date", ""))
        combo = f"{title}|{agency}|{date}".lower()
        combo_hash = hashlib.md5(combo.encode()).hexdigest()
        self.seen_hashes.add(combo_hash)

def make_session():
    """Create HTTP session with retry logic."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })

    retry = Retry(
        total=3,
        backoff_factor=1.0,
        status_forcelist=[403, 429, 500, 502, 503, 504]
    )
    session.mount("https://", HTTPAdapter(max_retries=retry))

    return session

def extract_detail_information(session, url):
    """Extract detailed information from opportunity page."""
    if not url:
        return {}

    try:
        response = session.get(url, timeout=TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        details = {}
        cleaner = DataCleaner()

        # Extract all tables for key-value pairs
        for table in soup.find_all('table'):
            for row in table.find_all('tr'):
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    label = cleaner.clean_text(cells[0].get_text()).lower()
                    value = cleaner.clean_text(cells[1].get_text())

                    if label and value:
                        # Store all details with cleaned labels
                        details[label] = value

        # Extract specific patterns from full text
        full_text = soup.get_text()

        # Look for email addresses
        emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', full_text)
        if emails:
            details['extracted_emails'] = ', '.join(set(emails[:3]))  # Keep first 3 unique

        # Look for phone numbers
        phones = re.findall(r'(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}', full_text)
        if phones:
            details['extracted_phones'] = ', '.join(set(phones[:3]))

        # Look for dollar amounts
        amounts = re.findall(r'\$[\d,]+(?:\.\d{2})?(?:\s*(?:million|billion|M|B))?', full_text)
        if amounts:
            details['estimated_values'] = ', '.join(amounts[:3])

        # Look for attachments/documents
        attachments = []
        for link in soup.find_all('a', href=True):
            href = link['href'].lower()
            if any(ext in href for ext in ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.zip']):
                attachments.append(cleaner.clean_text(link.get_text()))

        if attachments:
            details['attachments'] = f"{len(attachments)} documents: {', '.join(attachments[:5])}"

        # Look for important dates
        date_patterns = [
            (r'pre[\s-]?bid', 'pre-bid conference'),
            (r'q(?:uestion)?[\s&]*a(?:nswer)?', 'q&a deadline'),
            (r'amendment', 'amendment date'),
            (r'walk[\s-]?through', 'walkthrough date'),
        ]

        for pattern, label in date_patterns:
            matches = re.finditer(pattern + r'.*?(\d{1,2}/\d{1,2}/\d{4})', full_text, re.IGNORECASE)
            for match in matches:
                details[label] = match.group(1)
                break

        return details

    except Exception as e:
        logger.debug(f"Failed to fetch details from {url}: {e}")
        return {}

def extract_records_from_page(soup, session):
    """Extract records from the listing page."""
    records = []
    cleaner = DataCleaner()

    # Find the main table
    table = soup.find("table", class_="iv-grid-view")
    if not table:
        logger.warning("No results table found")
        return records

    # Get all rows (skip header)
    rows = table.find_all("tr")[1:]

    for row in rows:
        cells = row.find_all("td")
        if len(cells) < 8:
            continue

        record = {}

        # Basic extraction from listing
        # Column mapping based on actual eMMA structure:
        # 0: Edit link, 1: ID, 2: Title, 3: Status, 4: Due Date, 5: Publish Date,
        # 6: Category, 7: Type, 8: Agency, 12: Officer, 14: Sub-agency, 17: eMMA ID

        # Solicitation ID
        if len(cells) > 1:
            record["solicitation_number"] = cleaner.clean_text(cells[1].get_text())

        # Title and URL
        if len(cells) > 2:
            title_cell = cells[2]
            record["opportunity_title"] = cleaner.clean_text(title_cell.get_text())

            # Get URL from edit link
            if len(cells) > 0:
                edit_cell = cells[0]
                link = edit_cell.find("a")
                if link and link.get("href"):
                    record["opportunity_url"] = urljoin(BASE, link["href"])

        # Status
        if len(cells) > 3:
            record["status"] = cleaner.clean_text(cells[3].get_text())

        # Due Date
        if len(cells) > 4:
            due_text = cleaner.clean_text(cells[4].get_text())
            record["response_deadline"] = due_text
            parsed_due = cleaner.parse_date(due_text)
            if parsed_due:
                days_until = (parsed_due - datetime.now()).days
                record["days_until_due"] = days_until

        # Publish Date
        if len(cells) > 5:
            pub_text = cleaner.clean_text(cells[5].get_text())
            record["published_date"] = pub_text

        # Category
        if len(cells) > 6:
            record["category"] = cleaner.clean_text(cells[6].get_text())

        # Solicitation Type
        if len(cells) > 7:
            record["procurement_type"] = cleaner.clean_text(cells[7].get_text())

        # Issuing Agency
        if len(cells) > 8:
            record["issuing_agency"] = cleaner.clean_text(cells[8].get_text())

        # Procurement Officer
        if len(cells) > 12:
            officer = cleaner.clean_text(cells[12].get_text())
            if officer:
                record["procurement_officer"] = officer

        # Build additional information from extra columns
        additional_info = {}

        # Sub-agency
        if len(cells) > 14:
            sub_agency = cleaner.clean_text(cells[14].get_text())
            if sub_agency:
                additional_info["sub_agency"] = sub_agency

        # eMMA System ID
        if len(cells) > 17:
            emma_id = cleaner.clean_text(cells[17].get_text())
            if emma_id:
                additional_info["emma_system_id"] = emma_id

        # Auto-opening info
        if len(cells) > 9:
            auto_opening = cleaner.clean_text(cells[9].get_text())
            if auto_opening:
                additional_info["auto_opening"] = auto_opening

        # Round number
        if len(cells) > 10:
            round_num = cleaner.clean_text(cells[10].get_text())
            if round_num:
                additional_info["round_number"] = round_num

        # Award status
        if len(cells) > 11:
            award_status = cleaner.clean_text(cells[11].get_text())
            if award_status:
                additional_info["award_status"] = award_status

        # Authority
        if len(cells) > 13:
            authority = cleaner.clean_text(cells[13].get_text())
            if authority:
                additional_info["contracting_authority"] = authority

        # Site
        if len(cells) > 15:
            site = cleaner.clean_text(cells[15].get_text())
            if site:
                additional_info["site_location"] = site

        # Fetch detail page for more information
        if record.get("opportunity_url"):
            detail_info = extract_detail_information(session, record["opportunity_url"])

            # Extract contact info if found
            if detail_info:
                # Look for email
                for key in ['email', 'contact email', 'buyer email', 'extracted_emails']:
                    if key in detail_info:
                        email = cleaner.validate_email(detail_info[key].split(',')[0])
                        if email:
                            record["contact_email"] = email
                            break

                # Look for phone
                for key in ['phone', 'telephone', 'contact number', 'extracted_phones']:
                    if key in detail_info:
                        phone = cleaner.validate_phone(detail_info[key].split(',')[0])
                        if phone:
                            record["contact_phone"] = phone
                            break

                # Add all other details to additional_info
                for key, value in detail_info.items():
                    if key not in ['email', 'phone', 'extracted_emails', 'extracted_phones']:
                        additional_info[key] = value

            time.sleep(SLEEP_BETWEEN)

        # Convert additional_info to JSON string for storage
        if additional_info:
            record["additional_information"] = json.dumps(additional_info, indent=2)
        else:
            record["additional_information"] = ""

        # Generate unique ID
        if record.get("solicitation_number"):
            record["unique_id"] = f"sol_{record['solicitation_number']}"
        else:
            combo = f"{record.get('opportunity_title', '')}|{record.get('issuing_agency', '')}"
            record["unique_id"] = f"hash_{hashlib.md5(combo.encode()).hexdigest()[:12]}"

        # Add metadata
        record["first_seen"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        record["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Calculate data quality score
        filled_fields = sum(1 for k, v in record.items()
                          if k != 'additional_information' and v and str(v).strip())
        total_fields = len([k for k in MAIN_COLUMNS if k != 'additional_information'])
        record["data_quality_score"] = round((filled_fields / total_fields) * 100, 1)

        records.append(record)

    return records

def scrape_emma():
    """Main scraping function."""
    session = make_session()
    duplicate_checker = DuplicateChecker()
    all_records = []

    try:
        # Get initial page
        response = session.get(BROWSE_URL, timeout=TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract records from first page
        records = extract_records_from_page(soup, session)

        for record in records:
            if not duplicate_checker.is_duplicate(record):
                duplicate_checker.mark_seen(record)
                all_records.append(record)
                logger.info(f"Extracted: {record.get('opportunity_title', 'Unknown')[:60]}...")
            else:
                logger.debug(f"Skipped duplicate: {record.get('opportunity_title', '')}")

        # Note: Pagination can be added here if needed

    except Exception as e:
        logger.error(f"Scraping failed: {e}")

    return all_records

def save_to_excel(records, filepath):
    """Save records to Excel with formatting."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Opportunities"

        # Write headers
        ws.append(MAIN_COLUMNS)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for record in records:
            row = []
            for col in MAIN_COLUMNS:
                value = record.get(col, "")
                row.append(value)
            ws.append(row)

        # Auto-size columns (limit width for additional_information)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    if column_letter == 'O':  # Additional information column
                        max_length = 50  # Limit width
                    else:
                        max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Wrap text for additional_information column
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter == 'O':  # Additional information column
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Add table formatting
        if ws.max_row > 1:
            tab = Table(displayName="OpportunitiesTable",
                       ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            style = TableStyleInfo(name="TableStyleMedium9",
                                  showFirstColumn=False,
                                  showLastColumn=False,
                                  showRowStripes=True,
                                  showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

        # Freeze header
        ws.freeze_panes = "A2"

        # Add summary sheet
        summary = wb.create_sheet("Summary")
        summary.append(["eMMA Opportunities Summary"])
        summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        summary.append([])
        summary.append(["Total Opportunities:", len(records)])
        summary.append([])

        # Analytics
        agencies = {}
        categories = {}
        urgent = 0

        for record in records:
            agency = record.get("issuing_agency", "Unknown")
            agencies[agency] = agencies.get(agency, 0) + 1

            category = record.get("category", "Unknown")
            categories[category] = categories.get(category, 0) + 1

            if record.get("days_until_due") is not None and record["days_until_due"] <= 7:
                urgent += 1

        summary.append([f"Urgent (Due within 7 days):", urgent])
        summary.append([])

        summary.append(["Top Agencies:"])
        for agency, count in sorted(agencies.items(), key=lambda x: x[1], reverse=True)[:5]:
            summary.append(["", agency, count])

        summary.append([])
        summary.append(["Top Categories:"])
        for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:5]:
            summary.append(["", category, count])

        # Save
        wb.save(filepath)
        logger.info(f"Saved {len(records)} records to {filepath}")

        return True

    except Exception as e:
        logger.error(f"Failed to save Excel: {e}")
        return False

def print_sample_data(records):
    """Print sample of extracted data with additional information."""
    print("\n" + "="*80)
    print("SAMPLE OF EXTRACTED DATA (WITH CONSOLIDATED ADDITIONAL INFO)")
    print("="*80)

    for i, record in enumerate(records[:2], 1):
        print(f"\nRecord {i}:")
        print("-" * 40)
        print(f"  Unique ID: {record.get('unique_id')}")
        print(f"  Solicitation #: {record.get('solicitation_number')}")
        print(f"  Title: {record.get('opportunity_title', '')[:70]}...")
        print(f"  Agency: {record.get('issuing_agency')}")
        print(f"  Category: {record.get('category')}")
        print(f"  Type: {record.get('procurement_type')}")
        print(f"  Status: {record.get('status')}")
        print(f"  Due Date: {record.get('response_deadline')} ({record.get('days_until_due')} days)")
        print(f"  Officer: {record.get('procurement_officer')}")
        print(f"  Email: {record.get('contact_email', 'Not found')}")
        print(f"  Phone: {record.get('contact_phone', 'Not found')}")
        print(f"  Data Quality: {record.get('data_quality_score')}%")

        # Show additional information
        additional = record.get('additional_information')
        if additional:
            try:
                info = json.loads(additional)
                if info:
                    print("\n  Additional Information:")
                    for key, value in list(info.items())[:5]:  # Show first 5 items
                        print(f"    - {key}: {str(value)[:60]}...")
            except:
                pass

    print("\n" + "="*80)
    print(f"TOTAL UNIQUE RECORDS: {len(records)}")
    print("="*80)

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="eMMA Scraper with Consolidated Info")
    parser.add_argument("--output", default=WORKBOOK_PATH, help="Output Excel file")

    args = parser.parse_args()

    logger.info("Starting eMMA scraping with consolidated columns...")

    # Scrape data
    records = scrape_emma()

    if records:
        # Print sample
        print_sample_data(records)

        # Save to Excel
        save_to_excel(records, args.output)

        print(f"\n✅ Successfully extracted {len(records)} unique opportunities")
        print(f"📁 Output saved to: {args.output}")
        print("\nThe 'additional_information' column contains all supplementary details in JSON format")
    else:
        logger.warning("No records extracted")

if __name__ == "__main__":
    main()